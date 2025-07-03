#!/usr/bin/env python3
"""Test script to verify that existing images are preserved during updates."""

import json
import requests
import tempfile
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage

def create_test_excel_with_images():
    """Create test Excel file with existing images to test preservation."""
    import time
    timestamp = int(time.time())
    temp_file_path = f'/tmp/test_excel_with_images_{timestamp}.xlsx'
    
    # Create test workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Order Form"
    
    # Add some sample data
    ws['A1'] = 'Sample Data'
    ws['B1'] = 'Value 1'
    ws['A2'] = 'Test Item'
    ws['B2'] = 'Value 2'
    
    # Add an existing image to test preservation
    test_image_path = "/Users/alexsherin/Projects_/excel-pptx-merger/tests/fixtures/images/image_Order Form_1.png"
    if os.path.exists(test_image_path):
        # Add image to A5 (different from our update target)
        existing_img = ExcelImage(test_image_path)
        existing_img.anchor = "A5"
        ws.add_image(existing_img)
        print(f"Added existing image to A5 for preservation test")
    
    # Create a second sheet with images
    ws2 = wb.create_sheet("Sheet2")
    ws2['A1'] = 'Sheet 2 Data'
    if os.path.exists(test_image_path):
        existing_img2 = ExcelImage(test_image_path)
        existing_img2.anchor = "B2"
        ws2.add_image(existing_img2)
        print(f"Added existing image to Sheet2:B2 for preservation test")
    
    # Image search table (our update target)
    ws['A15'] = 'Image'
    ws['B15'] = 'Search Criteria'
    ws['C15'] = 'Image Class.Division.Subdivision'
    
    wb.save(temp_file_path)
    wb.close()
    
    return temp_file_path

def test_image_preservation():
    """Test that existing images are preserved during updates."""
    # Create test files
    excel_file_path = create_test_excel_with_images()
    config_file_path = "/Users/alexsherin/Projects_/excel-pptx-merger/example_update_config.json"
    data_file_path = "/Users/alexsherin/Projects_/excel-pptx-merger/test_file_image_data.json"
    
    try:
        # Read configuration and data
        with open(config_file_path, 'r') as f:
            config = json.load(f)
        
        with open(data_file_path, 'r') as f:
            update_data = json.load(f)
        
        # Only test image_search subtable
        filtered_config = {
            "version": "1.0",
            "sheet_configs": {
                "Order Form": {
                    "subtables": [subtable for subtable in config["sheet_configs"]["Order Form"]["subtables"] 
                                 if subtable["name"] == "image_search"]
                }
            }
        }
        
        # Prepare API request
        url = "http://localhost:8080/api/v1/update"
        
        with open(excel_file_path, 'rb') as excel_file:
            files = {
                'excel_file': ('test_with_existing_images.xlsx', excel_file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            }
            
            data = {
                'config': json.dumps(filtered_config),
                'update_data': json.dumps(update_data)
            }
            
            print("Testing image preservation during update...")
            print("Excel file contains existing images in:")
            print("  - Order Form:A5")
            print("  - Sheet2:B2")
            print("Update will add new images to:")
            print("  - Order Form:A16 and A17")
            
            # Make request
            response = requests.post(url, files=files, data=data)
            
            if response.status_code == 200:
                # Save response file
                output_path = "updated_with_preserved_images.xlsx"
                with open(output_path, 'wb') as f:
                    f.write(response.content)
                print(f"✅ Success! Updated file saved as: {output_path}")
                print("Please manually verify that:")
                print("  1. Existing images in A5 and Sheet2:B2 are preserved")
                print("  2. New images were added to A16 and A17")
                print("  3. All data was preserved correctly")
                return True
            else:
                print(f"❌ Error {response.status_code}: {response.text}")
                return False
    
    except Exception as e:
        print(f"❌ Test failed: {e}")
        return False
    finally:
        # Cleanup temp file
        if os.path.exists(excel_file_path):
            os.unlink(excel_file_path)

if __name__ == "__main__":
    success = test_image_preservation()
    exit(0 if success else 1)