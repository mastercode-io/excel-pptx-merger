#!/usr/bin/env python3
"""Test script for the Excel Update API endpoint."""

import json
import requests
import tempfile
import os
from openpyxl import Workbook

def create_test_excel_file():
    """Create test Excel file for API testing."""
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_file.close()
    
    # Create test workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Order Form"
    
    # Add test data structure matching our config
    # Report info section
    ws['E12'] = 'Search/Audit or Monitoring & Representation'
    ws['F12'] = 'Original Report Title'
    
    # Client info section  
    ws['A2'] = 'Client'
    ws['B2'] = 'Original Client'
    ws['C2'] = 'Word Or Image'
    ws['D2'] = 'Word'
    
    # Search criteria table
    ws['A5'] = 'Word'
    ws['B5'] = 'Search Criteria'
    ws['C5'] = 'Remarks'
    ws['A6'] = 'OriginalWord'
    ws['B6'] = 'Exact'
    ws['C6'] = 'Old remark'
    
    # Image search table
    ws['A15'] = 'Image'
    ws['B15'] = 'Search Criteria'
    ws['C15'] = 'Image Class.Division.Subdivision'
    
    wb.save(temp_file.name)
    wb.close()
    
    return temp_file.name

def test_update_api():
    """Test the Excel update API endpoint."""
    # Create test files
    excel_file_path = create_test_excel_file()
    config_file_path = "/Users/alexsherin/Projects_/excel-pptx-merger/example_update_config.json"
    data_file_path = "/Users/alexsherin/Projects_/excel-pptx-merger/example_update_data.json"
    
    try:
        # Read configuration and data
        with open(config_file_path, 'r') as f:
            config = json.load(f)
        
        with open(data_file_path, 'r') as f:
            update_data = json.load(f)
        
        # Prepare API request
        url = "http://localhost:8080/api/v1/update"
        
        with open(excel_file_path, 'rb') as excel_file:
            files = {
                'excel_file': ('test_order_form.xlsx', excel_file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            }
            
            data = {
                'config': json.dumps(config),
                'update_data': json.dumps(update_data)
            }
            
            print("Sending API request to:", url)
            print("Config keys:", list(config.keys()))
            print("Update data keys:", list(update_data.keys()))
            
            # Make request
            response = requests.post(url, files=files, data=data)
            
            if response.status_code == 200:
                # Save response file
                output_path = "updated_test_file.xlsx"
                with open(output_path, 'wb') as f:
                    f.write(response.content)
                print(f"✅ Success! Updated file saved as: {output_path}")
                return True
            else:
                print(f"❌ Error {response.status_code}: {response.text}")
                return False
    
    except requests.exceptions.ConnectionError:
        print("❌ Connection error. Make sure the server is running on localhost:5000")
        print("Start server with: python scripts/run_local_server.py --debug")
        return False
    except Exception as e:
        print(f"❌ Test failed: {e}")
        return False
    finally:
        # Cleanup temp file
        if os.path.exists(excel_file_path):
            os.unlink(excel_file_path)

if __name__ == "__main__":
    print("Testing Excel Update API...")
    success = test_update_api()
    exit(0 if success else 1)