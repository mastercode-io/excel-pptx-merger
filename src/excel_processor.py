"""Excel data extraction and processing module with enhanced image position extraction."""

# CRITICAL STARTUP VERIFICATION - This should appear in server logs immediately
from datetime import datetime
print("=" * 80)
print("🚀 EXCEL_PROCESSOR.PY MODULE LOADING - UPDATED CODE VERSION")
print(f"📅 Load timestamp: {datetime.now()}")
print("=" * 80)

import logging
import os
import io
import uuid
import re
import base64
from typing import Any, Dict, List, Optional, Tuple, Union, BinaryIO
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker
from openpyxl.utils import get_column_letter, column_index_from_string
from PIL import Image as PILImage

from .temp_file_manager import TempFileManager
from .utils.exceptions import ExcelProcessingError, ValidationError
from .utils.validation import (
    normalize_column_name,
    validate_cell_range,
    is_empty_cell_value,
)
from .excel_range_exporter import ExcelRangeExporter, create_range_configs_from_dict
from .config_schema_validator import validate_config_file
from .utils.range_image_logger import (
    range_image_logger,
    setup_range_image_debug_mode,
    log_range_config,
    log_graph_api_status,
    log_range_export_progress,
    log_range_validation_result,
)

logger = logging.getLogger(__name__)


class ExcelProcessor:
    """Processes Excel files and extracts data according to configuration."""

    def __init__(
        self,
        file_input: Union[str, BinaryIO],
        graph_credentials: Optional[Dict[str, str]] = None,
    ) -> None:
        """Initialize Excel processor with file path or file-like object."""
        self.file_input = file_input
        self.file_path = None
        self.workbook = None
        self.data_frame = None
        self._is_memory_file = not isinstance(file_input, str)
        self._memory_file = None
        self._image_cache = {}
        self._range_exporter = None

        # Store Graph API credentials for later initialization
        self._graph_credentials = graph_credentials
        if not graph_credentials:
            range_image_logger.warning(
                "⚠️ No Graph API credentials provided - range image extraction disabled"
            )

        if self._is_memory_file:
            self._load_from_memory()
        else:
            self.file_path = file_input
            self._validate_file()

    def _validate_file(self) -> None:
        """Validate Excel file exists and is readable."""
        if not os.path.exists(self.file_path):
            raise ExcelProcessingError(f"Excel file not found: {self.file_path}")

        try:
            # Try to load workbook to validate format
            self.workbook = load_workbook(self.file_path, data_only=True)
            logger.info(f"Successfully loaded Excel file: {self.file_path}")
            
            # Initialize image cache for file-based workbooks too
            self._image_cache = {}
            # Pre-extract all image data immediately while file is accessible
            self._preload_image_data()
            
        except Exception as e:
            raise ExcelProcessingError(f"Invalid Excel file format: {e}")

    def _load_from_memory(self) -> None:
        """Load Excel file from memory (file-like object)."""
        try:
            # Ensure the file pointer is at the beginning
            if hasattr(self.file_input, "seek"):
                self.file_input.seek(0)

            # Read the entire file content into memory to avoid closed file issues
            # This ensures that image data can be accessed even if the original file handle closes
            import io

            file_content = self.file_input.read()
            if hasattr(self.file_input, "seek"):
                self.file_input.seek(0)  # Reset for any other operations

            # Create a new BytesIO object from the content
            memory_file = io.BytesIO(file_content)

            # Load workbook from the in-memory file
            self.workbook = load_workbook(memory_file, data_only=True)

            # Keep reference to the memory file to prevent it from being garbage collected
            self._memory_file = memory_file

            # Initialize image cache for pre-loaded data
            self._image_cache = {}

            # Pre-extract all image data immediately while file is accessible
            # This prevents issues with closed file handles later
            self._preload_image_data()

            logger.info("Successfully loaded Excel file from memory")
        except Exception as e:
            raise ExcelProcessingError(f"Invalid Excel file format from memory: {e}")

    def _preload_image_data(self) -> None:
        """Pre-extract all image data to prevent closed file handle issues."""
        logger.debug("Pre-loading image data to prevent file handle issues")
        image_count = 0
        error_count = 0

        for sheet_name in self.workbook.sheetnames:
            try:
                sheet = self.workbook[sheet_name]

                # Skip hidden sheets
                if sheet.sheet_state == "hidden":
                    logger.debug(f"Skipping hidden sheet: {sheet_name}")
                    continue

                # Check if sheet has images
                if not hasattr(sheet, "_images") or not sheet._images:
                    logger.debug(f"Sheet {sheet_name} has no images")
                    continue

                logger.debug(f"Pre-loading {len(sheet._images)} images from sheet: {sheet_name}")
                sheet_images = []
                
                for idx, image in enumerate(sheet._images):
                    try:
                        # Extract image data immediately while file handle is open
                        image_data = image._data()
                        if image_data:
                            # Cache the image data along with position info
                            cached_image = {
                                "data": image_data,
                                "anchor": (
                                    image.anchor if hasattr(image, "anchor") else None
                                ),
                                "original_image": image,  # Keep reference for position extraction
                            }
                            sheet_images.append(cached_image)
                            image_count += 1
                            logger.debug(f"Successfully cached image {idx} from sheet {sheet_name}")
                        else:
                            logger.warning(f"No data for image {idx} in sheet {sheet_name}")
                            # Add placeholder to maintain index consistency
                            sheet_images.append({
                                "data": None,
                                "anchor": None,
                                "original_image": image,
                                "error": "No image data available",
                            })
                            error_count += 1
                    except Exception as e:
                        logger.warning(
                            f"Failed to pre-load image {idx} from sheet {sheet_name}: {e}"
                        )
                        # Add placeholder to maintain index consistency
                        sheet_images.append({
                            "data": None,
                            "anchor": None,
                            "original_image": image,
                            "error": str(e),
                        })
                        error_count += 1

                # Always cache the sheet entry, even if empty, to avoid fallback to direct access
                self._image_cache[sheet_name] = sheet_images
                
            except Exception as e:
                logger.error(f"Error processing sheet {sheet_name} during pre-load: {e}")
                # Still add empty cache entry for this sheet to prevent direct file access
                self._image_cache[sheet_name] = []
                error_count += 1

        logger.info(
            f"Pre-loaded {image_count} images from {len(self._image_cache)} sheets "
            f"({error_count} errors encountered)"
        )

    def get_sheet_names(self) -> List[str]:
        """Get list of sheet names in the workbook."""
        if not self.workbook:
            self._validate_file()
        return list(self.workbook.sheetnames)

    def extract_data(
        self,
        global_settings: Dict[str, Any],
        sheet_config: Dict[str, Any],
        full_config: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        """Extract data from Excel sheet according to configuration.

        Args:
            global_settings: Global settings for extraction
            sheet_config: Configuration for each sheet
            full_config: Full configuration including range_images (optional)

        Returns:
            Dictionary of extracted data
        """
        try:
            # CRITICAL DEBUG: Entry point logging to confirm server is running updated code
            logger.info("=" * 80)
            logger.info("🚀 EXTRACT_DATA METHOD ENTRY POINT - SERVER CODE IS UPDATED")
            logger.info(f"📊 Global settings: {list(global_settings.keys()) if global_settings else 'None'}")
            logger.info(f"📊 Sheet config: {list(sheet_config.keys()) if sheet_config else 'None'}")
            logger.info(f"📊 Full config: {list(full_config.keys()) if full_config else 'None'}")
            logger.info("=" * 80)
            
            extracted_data = {}

            # Store global settings for use in other methods
            self._global_settings = global_settings

            # Extract images if image extraction is enabled
            all_images = {}
            if global_settings.get("image_extraction", {}).get("enabled", True):
                try:
                    # Extract images from all sheets
                    all_images = self.extract_images()
                    logger.debug(f"Extracted images from {len(all_images)} sheets")
                except Exception as e:
                    logger.warning(f"Failed to extract images: {e}")

            # Extract range images if configured and enabled
            range_images = {}
            logger.info("🔍 CHECKING RANGE IMAGE EXTRACTION CONDITIONS")
            # Check range image configuration
            range_images_in_config = (
                "range_images" in full_config if full_config else False
            )
            range_images_enabled = global_settings.get("range_images", {}).get(
                "enabled", True
            )

            logger.info(f"🔍 Full config exists: {full_config is not None}")
            logger.info(f"🔍 Range images in config: {range_images_in_config}")
            logger.info(f"🔍 Range images enabled: {range_images_enabled}")
            logger.info(f"🔍 Graph credentials available: {bool(self._graph_credentials)}")
            if full_config and "range_images" in full_config:
                logger.info(f"🔍 Number of range image configs: {len(full_config['range_images'])}")
            else:
                logger.info("🔍 No range_images key in full_config")

            if (
                full_config
                and "range_images" in full_config
                and global_settings.get("range_images", {}).get("enabled", True)
            ):
                # Enable debug logging for range extraction if configured
                if global_settings.get("range_images", {}).get("debug_logging", False):
                    range_image_logger.enable_debug()
                    logger.info("✅ Range extraction debug logging enabled")
                else:
                    range_image_logger.disable_debug()
                logger.info(
                    f"Starting range image extraction with {len(full_config['range_images'])} configurations"
                )
                try:
                    # Get SharePoint configuration from global settings
                    sharepoint_config = global_settings.get("sharepoint", {})
                    # Get tenant_id from config for Graph API
                    config_tenant_id = sharepoint_config.get("tenant_id")

                    # Update Graph API credentials with config tenant_id if available
                    if config_tenant_id and self._graph_credentials:
                        self._graph_credentials["tenant_id"] = config_tenant_id
                        logger.info("Using tenant_id from SharePoint configuration")

                    range_images = self._extract_range_images(
                        full_config["range_images"], sharepoint_config
                    )
                    logger.info(
                        f"Successfully extracted {len(range_images)} range images"
                    )
                except Exception as e:
                    logger.error(f"Failed to extract range images: {e}")
            else:
                logger.debug("Range image extraction skipped - conditions not met")

            for sheet_name, config in sheet_config.items():
                logger.debug(f"Processing sheet: {sheet_name}")
                
                # DEBUG gs_classes_terms sheet processing
                if sheet_name == "Order Form":
                    logger.info("⚡" * 50)
                    logger.info("⚡ ORDER FORM SHEET PROCESSING (contains gs_classes_terms)")
                    logger.info(f"⚡ Sheet config: {config}")
                    
                    # Check if this config contains gs_classes_terms subtable
                    subtables = config.get("subtables", [])
                    for subtable in subtables:
                        if subtable.get("name") == "gs_classes_terms":
                            logger.info(f"⚡ Found gs_classes_terms subtable config: {subtable}")
                            data_extraction = subtable.get("data_extraction", {})
                            column_mappings = data_extraction.get("column_mappings", {})
                            logger.info(f"⚡ gs_classes_terms column_mappings: {column_mappings}")
                    logger.info("⚡" * 50)

                if sheet_name not in self.workbook.sheetnames:
                    logger.warning(f"Sheet '{sheet_name}' not found in workbook")
                    continue

                worksheet = self.workbook[sheet_name]

                # Get images for this specific sheet
                sheet_images = all_images.get(sheet_name, [])

                sheet_data = self._process_sheet(worksheet, config, sheet_images)

                # Normalize sheet name for JSON compatibility
                normalized_sheet_name = normalize_column_name(sheet_name)
                logger.debug(
                    f"Normalized sheet name: '{sheet_name}' -> '{normalized_sheet_name}'"
                )

                extracted_data[normalized_sheet_name] = sheet_data

            # Add range images to extracted data if any were found
            if range_images:
                extracted_data["_range_images"] = range_images

            return extracted_data

        except Exception as e:
            raise ExcelProcessingError(f"Failed to extract data: {e}")

    def _process_sheet(
        self,
        worksheet: Worksheet,
        config: Dict[str, Any],
        images: Optional[List[Dict]] = None,
    ) -> Dict[str, Any]:
        """Process a single worksheet according to configuration."""
        sheet_data = {}

        if "subtables" not in config:
            raise ValidationError("Sheet configuration missing 'subtables'")

        for subtable_config in config["subtables"]:
            subtable_name = subtable_config.get("name", "unnamed_subtable")
            logger.debug(f"Processing subtable: {subtable_name}")

            try:
                subtable_data = self._extract_subtable(
                    worksheet, subtable_config, images
                )
                sheet_data[subtable_name] = subtable_data
            except Exception as e:
                logger.error(f"Failed to process subtable '{subtable_name}': {e}")
                sheet_data[subtable_name] = {}

        return sheet_data

    def _extract_subtable(
        self,
        worksheet: Worksheet,
        config: Dict[str, Any],
        images: Optional[List[Dict]] = None,
    ) -> Dict[str, Any]:
        """Extract data for a specific subtable configuration."""
        subtable_type = config.get("type", "table")
        subtable_name = config.get("name", "unnamed")
        header_search = config.get("header_search", {})
        data_extraction = config.get("data_extraction", {})
        
        # DEBUG ALL SUBTABLES
        logger.info(f"🔍 SUBTABLE: {subtable_name} (type: {subtable_type})")
        if subtable_name == "gs_classes_terms":
            logger.info("🔥" * 50)
            logger.info("🔥 GS_CLASSES_TERMS SUBTABLE PROCESSING")
            logger.info(f"🔥 Full subtable config: {config}")
            logger.info(f"🔥 data_extraction: {data_extraction}")
            logger.info(f"🔥 header_search: {header_search}")
            logger.info(f"🔥 subtable_type: {subtable_type}")
            
            # Check if data_extraction has column_mappings
            column_mappings_in_config = data_extraction.get("column_mappings", {})
            logger.info(f"🔥 column_mappings in data_extraction: {column_mappings_in_config}")
            
            # Log the exact keys and values in column_mappings
            for key, value in column_mappings_in_config.items():
                logger.info(f"🔥 Column mapping: '{key}' -> {value}")
            
            logger.info("🔥" * 50)

        # Find the header location
        header_location = self._find_header_location(worksheet, header_search)
        if not header_location:
            logger.warning(f"Header location not found for {subtable_name}")
            return {}
        else:
            if subtable_name == "gs_classes_terms":
                logger.info(f"🔍 GS_CLASSES_TERMS - header found at: {header_location}")

        # Extract data based on type
        if subtable_type == "key_value_pairs":
            return self._extract_key_value_pairs(
                worksheet, header_location, data_extraction, images
            )
        elif subtable_type == "table":
            return self._extract_table_data(
                worksheet, header_location, data_extraction, images
            )
        elif subtable_type == "matrix_table":
            return self._extract_matrix_table_data(
                worksheet, header_location, data_extraction, images
            )
        else:
            raise ValidationError(f"Unknown subtable type: {subtable_type}")

    def _find_header_location(
        self, worksheet: Worksheet, search_config: Dict[str, Any]
    ) -> Optional[Tuple[int, int]]:
        """Find header location based on search configuration."""
        method = search_config.get("method", "contains_text")
        search_text = search_config.get("text", "")
        search_column = search_config.get("column", "A")
        search_range = search_config.get("search_range", "A1:A10")

        if not validate_cell_range(search_range):
            raise ValidationError(f"Invalid cell range: {search_range}")

        try:
            if method == "contains_text":
                return self._find_by_text_contains(worksheet, search_text, search_range)
            elif method == "exact_match":
                return self._find_by_exact_match(worksheet, search_text, search_range)
            elif method == "regex":
                return self._find_by_regex(worksheet, search_text, search_range)
            else:
                raise ValidationError(f"Unknown search method: {method}")

        except Exception as e:
            logger.error(f"Header search failed: {e}")
            return None

    def _find_by_text_contains(
        self, worksheet: Worksheet, search_text: str, search_range: str
    ) -> Optional[Tuple[int, int]]:
        """Find cell containing specific text."""
        for row in worksheet[search_range]:
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if search_text.lower() in cell.value.lower():
                        return (cell.row, cell.column)
        return None

    def _find_by_exact_match(
        self, worksheet: Worksheet, search_text: str, search_range: str
    ) -> Optional[Tuple[int, int]]:
        """Find cell with exact text match."""
        for row in worksheet[search_range]:
            for cell in row:
                if cell.value and str(cell.value).strip() == search_text.strip():
                    return (cell.row, cell.column)
        return None

    def _find_by_regex(
        self, worksheet: Worksheet, pattern: str, search_range: str
    ) -> Optional[Tuple[int, int]]:
        """Find cell matching regex pattern."""
        try:
            regex = re.compile(pattern)
            for row in worksheet[search_range]:
                for cell in row:
                    if cell.value and regex.search(str(cell.value)):
                        return (cell.row, cell.column)
        except re.error as e:
            raise ValidationError(f"Invalid regex pattern: {e}")
        return None

    def _extract_key_value_pairs(
        self,
        worksheet: Worksheet,
        header_location: Tuple[int, int],
        config: Dict[str, Any],
        images: Optional[List[Dict]] = None,
    ) -> Dict[str, Any]:
        """Extract key-value pairs from Excel sheet."""
        header_row, header_col = header_location
        orientation = config.get("orientation", "horizontal")
        # Support for column offset - allows table to start in different column than search text
        headers_col_offset = config.get("headers_col_offset", 0)
        header_col = header_col + headers_col_offset
        max_pairs = (
            config.get("max_columns", 10)
            if orientation == "horizontal"
            else config.get("max_rows", 10)
        )
        column_mappings = config.get("column_mappings", {})

        data = {}
        field_types = {}  # Store field type information
        field_positions = {}  # Store cell position information for image fields

        try:
            if orientation == "horizontal":
                # Keys in one row, values in the next row
                keys_row = header_row + config.get("headers_row_offset", 0)
                values_row = header_row + config.get("data_row_offset", 1)

                for col_offset in range(max_pairs):
                    col = header_col + col_offset
                    key_cell = worksheet.cell(row=keys_row, column=col)
                    value_cell = worksheet.cell(row=values_row, column=col)

                    if key_cell.value and not is_empty_cell_value(key_cell.value):
                        original_key = str(key_cell.value).strip()
                        value = value_cell.value

                        # Apply column mapping if available
                        if original_key in column_mappings:
                            mapping = column_mappings[original_key]
                            if isinstance(mapping, str):
                                # Legacy format - just a string
                                mapped_key = mapping
                                field_type = "text"  # Default type
                            else:
                                # New format - object with name and type
                                mapped_key = mapping.get(
                                    "name", normalize_column_name(original_key)
                                )
                                field_type = mapping.get("type", "text")
                        else:
                            mapped_key = normalize_column_name(original_key)
                            field_type = "text"  # Default type

                        # Handle special field types
                        if field_type == "image":
                            # For image fields, we need to find the image at this position
                            # Store the exact cell position where this image field should be located
                            # Also store the original text value as fallback
                            data[mapped_key] = None
                            field_positions[mapped_key] = {
                                "row": values_row,
                                "col": col,
                                "text_fallback": value,
                            }
                        elif field_type == "link":
                            # For link fields, create structured data with title and link
                            data[mapped_key] = {
                                "title": original_key,  # Use field name as display text
                                "link": (
                                    str(value) if value is not None else ""
                                ),  # Use cell value as URL
                            }
                        else:
                            data[mapped_key] = value

                        # Store field type if not the default text type
                        if field_type != "text":
                            field_types[mapped_key] = field_type
            else:
                # Vertical orientation: keys in one column, values in adjacent column
                keys_col = header_col + config.get("headers_col_offset", 0)
                values_col = header_col + config.get("data_col_offset", 1)

                for row_offset in range(max_pairs):
                    row = header_row + config.get("headers_row_offset", 1) + row_offset
                    key_cell = worksheet.cell(row=row, column=keys_col)
                    value_cell = worksheet.cell(row=row, column=values_col)

                    if key_cell.value and not is_empty_cell_value(key_cell.value):
                        original_key = str(key_cell.value).strip()
                        value = value_cell.value

                        # Apply column mapping if available
                        if original_key in column_mappings:
                            mapping = column_mappings[original_key]
                            if isinstance(mapping, str):
                                # Legacy format - just a string
                                mapped_key = mapping
                                field_type = "text"  # Default type
                            else:
                                # New format - object with name and type
                                mapped_key = mapping.get(
                                    "name", normalize_column_name(original_key)
                                )
                                field_type = mapping.get("type", "text")
                        else:
                            mapped_key = normalize_column_name(original_key)
                            field_type = "text"  # Default type

                        # Handle special field types
                        if field_type == "image":
                            # For image fields, we need to find the image at this position
                            # Store the exact cell position where this image field should be located
                            # Also store the original text value as fallback
                            data[mapped_key] = None
                            field_positions[mapped_key] = {
                                "row": row,
                                "col": values_col,
                                "text_fallback": value,
                            }
                        elif field_type == "link":
                            # For link fields, create structured data with title and link
                            data[mapped_key] = {
                                "title": original_key,  # Use field name as display text
                                "link": (
                                    str(value) if value is not None else ""
                                ),  # Use cell value as URL
                            }
                        else:
                            data[mapped_key] = value

                        # Store field type if not the default text type
                        if field_type != "text":
                            field_types[mapped_key] = field_type

            # Add field type metadata if we have any non-text fields
            if field_types:
                data["_field_types"] = field_types

            # Add field position metadata if we have any image fields
            if field_positions:
                data["_field_positions"] = field_positions

            # Link images to image fields if images are available
            if images and data:
                self._link_images_to_key_value_pairs(data, images)

        except Exception as e:
            raise ExcelProcessingError(f"Failed to extract key-value pairs: {e}")

        return data

    def _extract_table_data(
        self,
        worksheet: Worksheet,
        header_location: Tuple[int, int],
        config: Dict[str, Any],
        images: Optional[List[Dict]] = None,
    ) -> List[Dict[str, Any]]:
        """Extract table data from Excel sheet."""
        header_row, header_col = header_location
        headers_row = header_row + config.get("headers_row_offset", 0)
        data_start_row = headers_row + config.get("data_row_offset", 1)
        # Support for column offset - allows table to start in different column than search text
        headers_col_offset = config.get("headers_col_offset", 0)
        header_col = header_col + headers_col_offset
        max_columns = config.get("max_columns", 20)
        max_rows = config.get("max_rows", 1000)
        column_mappings = config.get("column_mappings", {})
        # DEBUG ALL TABLES - Log every table extraction
        logger.info(f"🔍 TABLE EXTRACTION - column_mappings: {column_mappings}")
        logger.info(f"🔍 TABLE EXTRACTION - config keys: {list(config.keys())}")
        
        # SPECIAL DEBUG FOR gs_classes_terms TABLE
        is_gs_classes_terms = any("gs" in str(k).lower() and "class" in str(k).lower() for k in column_mappings.keys())
        if is_gs_classes_terms:
            logger.info("🎯" * 50)
            logger.info("🎯 GS_CLASSES_TERMS TABLE EXTRACTION DEBUG")
            logger.info(f"🎯 Full config passed to _extract_table_data: {config}")
            logger.info(f"🎯 Column mappings: {column_mappings}")
            logger.info(f"🎯 Column mapping keys: {list(column_mappings.keys())}")
            logger.info(f"🎯 Column mapping values: {list(column_mappings.values())}")
            for key, mapping in column_mappings.items():
                if isinstance(mapping, dict):
                    logger.info(f"🎯 Mapping '{key}' -> name: '{mapping.get('name')}', type: '{mapping.get('type')}'")
                else:
                    logger.info(f"🎯 Mapping '{key}' -> '{mapping}'")
            logger.info("🎯" * 50)

        try:
            # Extract headers
            headers = []
            original_headers = []  # Store original headers for mapping
            field_types = {}  # Store field types for each header

            for col_offset in range(max_columns):
                col = header_col + col_offset
                header_cell = worksheet.cell(row=headers_row, column=col)

                if header_cell.value and not is_empty_cell_value(header_cell.value):
                    header = str(header_cell.value).strip()
                    original_headers.append(header)

                    # Apply column mapping if available
                    if header in column_mappings:
                        mapping = column_mappings[header]
                        if isinstance(mapping, str):
                            # Legacy format - just a string
                            mapped_header = mapping
                            field_type = "text"  # Default type
                        else:
                            # New format - object with name and type
                            mapped_header = mapping.get(
                                "name", normalize_column_name(header)
                            )
                            field_type = mapping.get("type", "text")
                        
                        # DEBUG gs_classes_terms header mapping
                        if is_gs_classes_terms:
                            logger.info(f"🎯 HEADER MAPPING: '{header}' found in column_mappings")
                            logger.info(f"🎯 HEADER MAPPING: mapping = {mapping}")
                            logger.info(f"🎯 HEADER MAPPING: mapped_header = '{mapped_header}'")
                            logger.info(f"🎯 HEADER MAPPING: field_type = '{field_type}'")
                    else:
                        mapped_header = normalize_column_name(header)
                        field_type = "text"  # Default type
                        
                        # DEBUG gs_classes_terms header mapping
                        if is_gs_classes_terms:
                            logger.info(f"🎯 HEADER NORMALIZATION: '{header}' NOT found in column_mappings")
                            logger.info(f"🎯 HEADER NORMALIZATION: normalized to '{mapped_header}'")

                    headers.append(mapped_header)
                    field_types[mapped_header] = field_type
                else:
                    break  # Stop when we hit an empty header

            if not headers:
                logger.warning("No headers found for table extraction")
                return []

            # Extract data rows
            data_rows = []
            consecutive_empty_rows = 0

            for row_offset in range(max_rows):
                row = data_start_row + row_offset
                row_data = {}
                has_data = False

                for col_offset, (header, original_header) in enumerate(
                    zip(headers, original_headers)
                ):
                    col = header_col + col_offset
                    cell = worksheet.cell(row=row, column=col)
                    value = cell.value

                    # Check for image at this cell position
                    image_data = None
                    if images:
                        image_data = self._get_image_at_position(row, col, images)

                    # Handle mixed content (text + image), image only, or text only
                    if image_data and not is_empty_cell_value(value):
                        # Mixed content: both text and image
                        cell_value = {"text": value, "base64": image_data["base64"]}
                        if "path" in image_data:
                            cell_value["path"] = image_data["path"]
                        has_data = True
                    elif image_data:
                        # Image only
                        cell_value = image_data
                        has_data = True
                    else:
                        # Check for potential cell-embedded images (not detectable by openpyxl)
                        cell_embedded_info = self._check_for_cell_embedded_image(
                            cell, row, col
                        )

                        if cell_embedded_info:
                            # Cell likely contains embedded image that we can't extract
                            cell_value = cell_embedded_info
                            has_data = True
                        else:
                            # Text only (or empty)
                            cell_value = value
                            if not is_empty_cell_value(value):
                                has_data = True

                    # Use the mapped header name for the key
                    row_data[header] = cell_value

                    # Store field type as metadata if not the default text type
                    field_type = field_types.get(header, "text")
                    if field_type != "text":
                        if "_field_types" not in row_data:
                            row_data["_field_types"] = {}
                        row_data["_field_types"][header] = field_type

                if not has_data:
                    consecutive_empty_rows += 1
                    # Stop extraction after 2+ consecutive empty rows
                    if consecutive_empty_rows > 2:
                        break
                    # Skip this empty row but continue processing
                    continue
                else:
                    consecutive_empty_rows = 0  # Reset counter when we find data

                # Only add non-empty rows with actual data (skip header row data)
                if has_data and row > headers_row:
                    data_rows.append(row_data)

            return data_rows

        except Exception as e:
            raise ExcelProcessingError(f"Failed to extract table data: {e}")

    def _extract_matrix_table_data(
        self,
        worksheet: Worksheet,
        header_location: Tuple[int, int],
        config: Dict[str, Any],
        images: Optional[List[Dict]] = None,
    ) -> Dict[str, Dict[str, Any]]:
        """Extract matrix table data with row keys and column headers.

        Returns nested dictionary: {row_key: {col_key: value}}
        """
        header_row, header_col = header_location
        headers_row = header_row + config.get("headers_row_offset", 0)
        data_start_row = headers_row + config.get("data_row_offset", 1)

        # Support for column offset - allows table to start in different column than search text
        headers_col_offset = config.get("headers_col_offset", 0)
        header_col = header_col + headers_col_offset

        # Row keys are in the first column of the data area
        row_keys_col_offset = config.get("row_keys_col_offset", 0)
        row_keys_col = header_col + row_keys_col_offset

        # Data starts in the column after row keys
        data_col_offset = config.get("data_col_offset", 1)
        data_start_col = row_keys_col + data_col_offset

        max_columns = config.get("max_columns", 20)
        max_rows = config.get("max_rows", 1000)
        column_mappings = config.get("column_mappings", {})
        row_key_mappings = config.get("row_key_mappings", {})

        try:
            # Extract column headers
            headers = []
            original_headers = []
            field_types = {}

            for col_offset in range(max_columns):
                col = data_start_col + col_offset
                header_cell = worksheet.cell(row=headers_row, column=col)

                if header_cell.value and not is_empty_cell_value(header_cell.value):
                    header = str(header_cell.value).strip()
                    original_headers.append(header)
                    # DEBUG ALL TABLES - Log every header found
                    logger.info(f"🔍 HEADER FOUND: '{header}'")

                    # Apply column mapping if available
                    if header in column_mappings:
                        mapping = column_mappings[header]
                        if isinstance(mapping, str):
                            mapped_header = mapping
                            field_type = "text"
                        else:
                            mapped_header = mapping.get(
                                "name", normalize_column_name(header)
                            )
                            field_type = mapping.get("type", "text")
                        # DEBUG ALL TABLES - Log every mapping applied
                        logger.info(f"🔍 MAPPING APPLIED: '{header}' -> '{mapped_header}'")
                    else:
                        mapped_header = normalize_column_name(header)
                        field_type = "text"
                        # DEBUG ALL TABLES - Log every normalization fallback
                        logger.info(f"🔍 NO MAPPING: '{header}' -> normalized: '{mapped_header}'")

                    headers.append(mapped_header)
                    field_types[mapped_header] = field_type
                else:
                    break

            if not headers:
                logger.warning("No column headers found for matrix table extraction")
                return {}

            # Extract matrix data
            matrix_data = {}
            consecutive_empty_rows = 0

            for row_offset in range(max_rows):
                row = data_start_row + row_offset

                # Get row key from first column
                row_key_cell = worksheet.cell(row=row, column=row_keys_col)

                if not row_key_cell.value or is_empty_cell_value(row_key_cell.value):
                    consecutive_empty_rows += 1
                    if (
                        consecutive_empty_rows >= 3
                    ):  # Stop after 3 consecutive empty rows
                        break
                    continue

                consecutive_empty_rows = 0
                original_row_key = str(row_key_cell.value).strip()

                # Apply row key mapping if available
                if original_row_key in row_key_mappings:
                    mapped_row_key = row_key_mappings[original_row_key]
                else:
                    mapped_row_key = normalize_column_name(original_row_key)

                # Extract row data
                row_data = {}
                has_data = False

                for col_offset, (header, original_header) in enumerate(
                    zip(headers, original_headers)
                ):
                    col = data_start_col + col_offset
                    cell = worksheet.cell(row=row, column=col)
                    value = cell.value

                    # Check for image at this cell position
                    image_data = None
                    if images:
                        image_data = self._get_image_at_position(row, col, images)

                    # Handle mixed content (text + image), image only, or text only
                    if image_data and not is_empty_cell_value(value):
                        # Mixed content: both text and image
                        cell_value = {"text": value, "base64": image_data["base64"]}
                        if "path" in image_data:
                            cell_value["path"] = image_data["path"]
                        has_data = True
                    elif image_data:
                        # Image only
                        cell_value = image_data
                        has_data = True
                    else:
                        # Check for potential cell-embedded images
                        cell_embedded_info = self._check_for_cell_embedded_image(
                            cell, row, col
                        )
                        if cell_embedded_info:
                            cell_value = cell_embedded_info
                            has_data = True
                        else:
                            # Regular text/value
                            cell_value = value
                            if not is_empty_cell_value(value):
                                has_data = True

                    row_data[header] = cell_value

                # Only add rows with actual data
                if has_data:
                    matrix_data[mapped_row_key] = row_data

            # Add field types metadata
            if field_types:
                matrix_data["_field_types"] = field_types

            return matrix_data

        except Exception as e:
            raise ExcelProcessingError(f"Failed to extract matrix table data: {e}")

    def extract_images(
        self, session_dir: Optional[str] = None
    ) -> Dict[str, List[Dict[str, Any]]]:
        """Extract images from Excel file and optionally save them to disk.

        Args:
            session_dir: Optional session directory for saving images. If None and
                        SAVE_FILES=False, images will only be encoded as base64.

        Returns:
            Dictionary mapping sheet names to lists of image metadata.
            Each image metadata includes format, position, base64 encoding, and
            optionally path/filename if saved to disk.
        """
        # Get app configuration to check SAVE_FILES setting
        from .config_manager import ConfigManager

        config_manager = ConfigManager()
        app_config = config_manager.get_app_config()
        save_files = app_config.get("save_files", False)

        logger.info(f"Extracting images from Excel file - save_files: {save_files}")

        # Initialize temp file manager if we need to save files
        temp_manager = None
        if save_files and session_dir:
            temp_manager = TempFileManager()
            logger.info(f"Using session directory for images: {session_dir}")
        elif save_files and self.file_path:
            # Fallback to file path directory if available
            file_dir = os.path.dirname(self.file_path)
            if "input" in file_dir:
                session_dir = os.path.dirname(file_dir)
            else:
                session_dir = file_dir
            temp_manager = TempFileManager()
            logger.info(f"Using inferred session directory for images: {session_dir}")

        # Initialize result dictionary
        result = {}

        # Process each worksheet
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]

            # Skip hidden sheets
            if sheet.sheet_state == "hidden":
                logger.debug(f"Skipping hidden sheet: {sheet_name}")
                continue

            # Initialize list for this sheet's images
            result[sheet_name] = []

            # Check if sheet has images
            if not hasattr(sheet, "_images"):
                logger.debug(f"Sheet {sheet_name} has no images")
                continue

            logger.info(
                f"Processing {len(sheet._images)} images in sheet: {sheet_name}"
            )

            # Process each image in the sheet
            for idx, image in enumerate(sheet._images):
                try:
                    # Get image data from cache if available (memory loading) or directly (file loading)
                    image_data = None
                    cached_image = None

                    # Always try cache first, regardless of file type
                    if (
                        hasattr(self, "_image_cache")
                        and sheet_name in self._image_cache
                        and idx < len(self._image_cache[sheet_name])
                    ):
                        # Use cached data (available for all workbook types now)
                        cached_image = self._image_cache[sheet_name][idx]
                        image_data = cached_image.get("data")
                        if cached_image.get("error"):
                            logger.warning(
                                f"Cannot extract image {idx} from sheet {sheet_name}: {cached_image['error']}. Skipping this image."
                            )
                            continue
                        elif not image_data:
                            logger.warning(
                                f"No cached data for image {idx} in sheet {sheet_name}. Skipping this image."
                            )
                            continue
                    else:
                        # Fallback to direct access with improved error handling
                        try:
                            logger.debug(f"Attempting direct image data access for image {idx} in sheet {sheet_name}")
                            image_data = image._data()
                        except ValueError as e:
                            if "I/O operation on closed file" in str(e):
                                logger.error(
                                    f"File handle closed for image {idx} in sheet {sheet_name}. "
                                    f"This suggests the workbook file was closed prematurely. Skipping this image."
                                )
                            else:
                                logger.warning(
                                    f"ValueError accessing image {idx} from sheet {sheet_name}: {e}. Skipping this image."
                                )
                            continue
                        except Exception as e:
                            logger.warning(
                                f"Cannot extract image {idx} from sheet {sheet_name}: {type(e).__name__}: {e}. Skipping this image."
                            )
                            continue

                    if not image_data:
                        logger.warning(f"No data for image {idx} in sheet {sheet_name}")
                        continue

                    # Determine image format
                    img_format = self._detect_image_format(image_data)
                    if not img_format:
                        logger.warning(
                            f"Could not detect format for image {idx} in sheet {sheet_name}"
                        )
                        img_format = "png"  # Default to PNG

                    # Generate a unique filename
                    image_filename = f"image_{sheet_name}_{idx}.{img_format}"

                    # Extract position information (use cached image if available)
                    position_source = (
                        cached_image.get("original_image") if cached_image else image
                    )
                    position = self._extract_image_position(position_source)

                    # Encode image as base64
                    image_base64 = self._encode_image_as_base64(image_data, img_format)

                    # Initialize image metadata
                    image_meta = {
                        "filename": image_filename,
                        "sheet": sheet_name,
                        "index": idx,
                        "format": img_format,
                        "position": position,
                        "image_base64": image_base64,
                    }

                    # Conditionally save image to disk
                    if save_files and temp_manager and session_dir:
                        try:
                            image_path = temp_manager.save_file_to_temp(
                                session_dir,
                                image_filename,
                                image_data,
                                temp_manager.FILE_TYPE_IMAGE,
                            )

                            # Verify the image was saved
                            if os.path.exists(image_path):
                                image_meta["path"] = image_path
                                logger.debug(f"Saved image to disk: {image_path}")
                            else:
                                logger.warning(
                                    f"Failed to save image {idx} from sheet {sheet_name}"
                                )
                        except Exception as e:
                            logger.warning(f"Error saving image to disk: {e}")
                    else:
                        # Not saving files, only using base64
                        logger.debug(
                            f"Image {idx} from sheet {sheet_name} processed in-memory only"
                        )

                    # Add to results
                    result[sheet_name].append(image_meta)
                    log_msg = (
                        f"Extracted image: {image_meta.get('path', 'in-memory only')}"
                    )
                    logger.debug(log_msg)

                except Exception as e:
                    logger.error(
                        f"Error extracting image {idx} from sheet {sheet_name}: {e}"
                    )
                    logger.exception("Image extraction error details")

            logger.info(
                f"Extracted {len(result[sheet_name])} images from sheet {sheet_name}"
            )

        # Log summary
        total_images = sum(len(images) for images in result.values())
        logger.info(f"Total images extracted: {total_images}")

        return result

    def get_image_by_position(
        self,
        images: Dict[str, List[Dict[str, Any]]],
        target_cell: str,
        sheet_name: Optional[str] = None,
    ) -> Optional[Dict[str, Any]]:
        """Find image by cell position or proximity."""
        sheets_to_search = [sheet_name] if sheet_name else images.keys()

        for sheet in sheets_to_search:
            if sheet not in images:
                continue

            for image_info in images[sheet]:
                position = image_info.get("position", {})

                # Exact match on from_cell or estimated_cell
                if (
                    position.get("from_cell") == target_cell
                    or position.get("estimated_cell") == target_cell
                ):
                    return image_info

                # Check if target position falls within image range
                if position.get("coordinates"):
                    coords = position["coordinates"]
                    if "from" in coords and "to" in coords:
                        target_coords = self._cell_to_coordinates(target_cell)
                        if target_coords and self._is_in_range(target_coords, coords):
                            return image_info

        return None

    def _cell_to_coordinates(self, cell_ref: str) -> Optional[Dict[str, int]]:
        """Convert cell reference like 'A1' to coordinates."""
        try:
            match = re.match(r"^([A-Z]+)(\d+)$", cell_ref.upper())
            if match:
                col_str, row_str = match.groups()
                return {
                    "col": column_index_from_string(col_str) - 1,  # 0-based
                    "row": int(row_str) - 1,  # 0-based
                }
        except Exception as e:
            logger.debug(f"Failed to convert cell reference {cell_ref}: {e}")
        return None

    def _is_in_range(self, target: Dict[str, int], range_coords: Dict) -> bool:
        """Check if target coordinates fall within the given range."""
        try:
            from_coords = range_coords["from"]
            to_coords = range_coords["to"]

            return (
                from_coords["col"] <= target["col"] <= to_coords["col"]
                and from_coords["row"] <= target["row"] <= to_coords["row"]
            )
        except (KeyError, TypeError):
            return False

    def get_cell_value(self, sheet_name: str, cell_reference: str) -> Any:
        """Get value from specific cell."""
        if not self.workbook:
            self._validate_file()

        try:
            if sheet_name not in self.workbook.sheetnames:
                raise ValidationError(f"Sheet '{sheet_name}' not found")

            worksheet = self.workbook[sheet_name]
            cell = worksheet[cell_reference]
            return cell.value

        except Exception as e:
            raise ExcelProcessingError(f"Failed to get cell value: {e}")

    def get_range_values(self, sheet_name: str, cell_range: str) -> List[List[Any]]:
        """Get values from cell range."""
        if not self.workbook:
            self._validate_file()

        try:
            if sheet_name not in self.workbook.sheetnames:
                raise ValidationError(f"Sheet '{sheet_name}' not found")

            if not validate_cell_range(cell_range):
                raise ValidationError(f"Invalid cell range: {cell_range}")

            worksheet = self.workbook[sheet_name]
            cell_range_obj = worksheet[cell_range]

            # Handle single cell vs range
            if hasattr(cell_range_obj, "__iter__") and not isinstance(
                cell_range_obj, str
            ):
                return [[cell.value for cell in row] for row in cell_range_obj]
            else:
                return [[cell_range_obj.value]]

        except Exception as e:
            raise ExcelProcessingError(f"Failed to get range values: {e}")

    def to_dataframe(
        self, sheet_name: str, header_row: Optional[int] = None
    ) -> pd.DataFrame:
        """Convert Excel sheet to pandas DataFrame."""
        try:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=header_row)
            return df

        except Exception as e:
            raise ExcelProcessingError(f"Failed to convert to DataFrame: {e}")

    def get_image_summary(
        self, images: Dict[str, List[Dict[str, Any]]]
    ) -> Dict[str, Any]:
        """Get summary information about extracted images."""
        summary = {
            "total_images": 0,
            "sheets_with_images": 0,
            "images_by_sheet": {},
            "position_summary": {
                "with_position": 0,
                "without_position": 0,
                "position_types": {},
            },
        }

        for sheet_name, sheet_images in images.items():
            if sheet_images:
                summary["sheets_with_images"] += 1
                summary["images_by_sheet"][sheet_name] = len(sheet_images)
                summary["total_images"] += len(sheet_images)

                for image_info in sheet_images:
                    position = image_info.get("position", {})
                    if position.get("estimated_cell"):
                        summary["position_summary"]["with_position"] += 1
                        anchor_type = position.get("anchor_type", "unknown")
                        summary["position_summary"]["position_types"][anchor_type] = (
                            summary["position_summary"]["position_types"].get(
                                anchor_type, 0
                            )
                            + 1
                        )
                    else:
                        summary["position_summary"]["without_position"] += 1

        return summary

    def link_images_to_table(
        self, extracted_data: Dict[str, Any], images: Dict[str, List[Dict[str, Any]]]
    ) -> Dict[str, Any]:
        """Link extracted images to their corresponding rows in the image search table.

        This method maps images to the image search table rows based on their position
        in the Excel sheet, replacing the image cell value with the path to the extracted image.

        Args:
            extracted_data: The data extracted from Excel sheets
            images: Dictionary of extracted images with position information

        Returns:
            Updated extracted data with image paths linked to table rows
        """
        if not images:
            return extracted_data

        # Create a deep copy to avoid modifying the original data
        import copy

        result = copy.deepcopy(extracted_data)

        # Process each sheet
        for sheet_name, sheet_data in result.items():
            # Get the original sheet name (before normalization)
            original_sheet_names = [
                name
                for name in images.keys()
                if normalize_column_name(name) == sheet_name
            ]

            if not original_sheet_names:
                logger.debug(
                    f"No matching original sheet name found for normalized name '{sheet_name}'"
                )
                continue

            original_sheet_name = original_sheet_names[0]
            sheet_images = images.get(original_sheet_name, [])

            logger.debug(
                f"Found original sheet name '{original_sheet_name}' for normalized name '{sheet_name}'"
            )

            if not sheet_images:
                logger.debug(f"No images found for sheet '{original_sheet_name}'")
                continue

            # Check for image_search table
            if "image_search" in sheet_data and isinstance(
                sheet_data["image_search"], list
            ):
                image_search_table = sheet_data["image_search"]

                # Map row numbers to images based on position
                row_to_image = {}
                for img in sheet_images:
                    if "position" in img and "coordinates" in img["position"]:
                        row_num = img["position"]["coordinates"]["from"]["row"]

                        # Always include base64 data in the image reference
                        image_data = {
                            "base64": img["image_base64"],  # Always include base64 data
                        }

                        # Include path if it exists (for debugging/logging)
                        if "path" in img and os.path.exists(img["path"]):
                            image_data["path"] = img["path"]

                        row_to_image[row_num] = image_data
                        logger.debug(
                            f"Mapped image at row {row_num} to data with base64 content"
                        )

                # Sort images by row number
                sorted_rows = sorted(row_to_image.keys())

                # Link images to table rows
                for i, row_data in enumerate(image_search_table):
                    if i < len(sorted_rows):
                        row_data["image"] = row_to_image[sorted_rows[i]]
                        logger.debug(
                            f"Linked image at Excel row {sorted_rows[i]} to table row {i}"
                        )

        return result

    def extract_single_sheet(
        self,
        sheet_name: str,
        config: Optional[Dict] = None,
        auto_detect: bool = True,
        max_rows: Optional[int] = None,
    ) -> Dict[str, Any]:
        """Extract data from a single specified sheet with auto-detection support.

        Args:
            sheet_name: Name of the Excel sheet to process
            config: Optional extraction configuration
            auto_detect: Whether to auto-detect structure when no config provided
            max_rows: Optional maximum number of rows to extract

        Returns:
            Dictionary containing extracted data and metadata
        """
        # Validate sheet exists
        if sheet_name not in self.workbook.sheetnames:
            available_sheets = ", ".join(self.workbook.sheetnames)
            raise ValidationError(
                f"Sheet '{sheet_name}' not found. Available sheets: {available_sheets}"
            )

        worksheet = self.workbook[sheet_name]

        # Create or use provided configuration
        if config is None and auto_detect:
            logger.info(f"Auto-detecting structure for sheet: {sheet_name}")
            # If max_rows is None, we want to scan all rows
            scan_all_rows = max_rows is None
            detection_config = self._auto_detect_sheet_structure(
                worksheet, scan_all_rows=scan_all_rows
            )
            extraction_method = "auto_detect"
        elif config:
            detection_config = config
            extraction_method = "config_based"
        else:
            raise ValidationError("Either provide a config or enable auto_detect")

        # Apply max_rows limit to configuration if specified
        # If max_rows is None, ensure we extract all rows
        if max_rows is not None:
            self._apply_max_rows_to_config(detection_config, max_rows)
            logger.info(
                f"Applied max_rows limit of {max_rows} to extraction configuration"
            )
        else:
            # Ensure we extract all rows when max_rows is not specified
            self._apply_extract_all_rows(detection_config, worksheet.max_row)
            logger.info(f"Configured to extract all rows (up to {worksheet.max_row})")

        # Extract data using the configuration
        sheet_config = {sheet_name: detection_config}
        logger.debug(f"Using detection config for {sheet_name}: {detection_config}")
        extracted_data = self.extract_data({}, sheet_config)

        # Extract images from the sheet
        session_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "temp")
        os.makedirs(session_dir, exist_ok=True)
        images = self.extract_images(session_dir)

        # Link images to extracted data
        if images and sheet_name in images:
            normalized_sheet_name = normalize_column_name(sheet_name)
            sheet_data = extracted_data.get(normalized_sheet_name, {})

            # Find all subtables in the extracted data
            for subtable_name, subtable_data in sheet_data.items():
                if isinstance(subtable_data, list) and subtable_data:
                    # This is a table-like structure, link images to rows
                    self._link_images_to_rows(subtable_data, images[sheet_name])

        # Get sheet metadata
        metadata = self._get_sheet_metadata(
            worksheet,
            detection_config,
            extraction_method,
            extracted_data.get(normalize_column_name(sheet_name), {}),
        )

        return {
            "data": extracted_data.get(normalize_column_name(sheet_name), {}),
            "metadata": metadata,
        }

    def _apply_max_rows_to_config(self, config: Dict[str, Any], max_rows: int) -> None:
        """Apply max_rows limit to all subtables in the configuration."""
        if "subtables" not in config:
            return

        for subtable in config["subtables"]:
            if "data_extraction" in subtable:
                # Only apply to table type subtables (not key_value_pairs)
                if subtable.get("type", "table") == "table":
                    subtable["data_extraction"]["max_rows"] = max_rows

    def _apply_extract_all_rows(
        self, config: Dict[str, Any], max_sheet_rows: int
    ) -> None:
        """Configure extraction to include all rows in the sheet."""
        if "subtables" not in config:
            return

        for subtable in config["subtables"]:
            if "data_extraction" in subtable:
                # Only apply to table type subtables (not key_value_pairs)
                if subtable.get("type", "table") == "table":
                    # Set max_rows to the maximum number of rows in the sheet
                    # Add a buffer to ensure we get everything
                    subtable["data_extraction"]["max_rows"] = max_sheet_rows + 10

    def _get_image_at_position(
        self, row: int, col: int, images: List[Dict]
    ) -> Optional[Dict]:
        """Get image data if there's an image at the specified cell position.

        Args:
            row: 1-based row number from table extraction
            col: 1-based column number from table extraction
            images: List of image dictionaries with 0-based coordinates
        """
        for img in images:
            if "position" in img and "coordinates" in img["position"]:
                coords = img["position"]["coordinates"]["from"]
                # Convert 1-based table coordinates to 0-based image coordinates
                if coords["row"] == row - 1 and coords["col"] == col - 1:
                    image_data = {"base64": img["image_base64"]}
                    if "path" in img:
                        image_data["path"] = img["path"]
                    return image_data
        return None

    def _check_for_cell_embedded_image(
        self, cell, row: int, col: int
    ) -> Optional[Dict]:
        """Check if a cell likely contains an embedded image (detection only)."""
        from openpyxl.utils import get_column_letter

        cell_ref = f"{get_column_letter(col)}{row}"

        # Note: Complex extraction methods removed - detection only

        # Check for various indicators of embedded content
        indicators = {
            "has_hyperlink": cell.hyperlink is not None,
            "has_comment": cell.comment is not None,
            "data_type": cell.data_type,
            "number_format": cell.number_format,
            "fill_type": cell.fill.fill_type if cell.fill else None,
            "font_name": cell.font.name if cell.font else None,
        }

        # More specific detection for #VALUE! that might indicate images
        # Only treat as embedded image if it's specifically in keywords/images-related columns
        if isinstance(cell.value, str) and cell.value == "#VALUE!":
            # Only detect as embedded image in likely image columns (column A for keywords_images)
            if col == 1:  # Column A (keywords_images column)
                # Try to get more context about the cell
                return {
                    "type": "cell_embedded_image_placeholder",
                    "message": "Embedded image detected (extraction not supported)",
                    "cell_ref": cell_ref,
                    "original_value": cell.value,
                    "extraction_status": "requires_manual_export",
                    "note": "Cell-embedded images can be manually exported from Excel using 'Save as Web Page' or similar methods",
                    "suggested_workflow": "Export Excel as HTML or use specialized tools to extract embedded images",
                }
            else:
                # For #VALUE! in other columns, treat as formula error, not image
                return None

        # Check for cells that might contain "Picture" or similar indicators
        if isinstance(cell.value, str) and cell.value.lower() in ["picture", "image"]:
            return {
                "type": "cell_embedded_image_placeholder",
                "message": f"Cell contains '{cell.value}' indicating embedded image",
                "cell_ref": cell_ref,
                "original_value": cell.value,
                "extraction_status": "not_available",
            }

        return None

    def _link_images_to_rows(
        self, rows: List[Dict[str, Any]], images: List[Dict[str, Any]]
    ) -> None:
        """Link images to data rows based on position information."""
        if not images or not rows:
            return

        # Create a mapping of row numbers to images
        row_to_image = {}
        for img in images:
            if "position" in img and "coordinates" in img["position"]:
                row_num = img["position"]["coordinates"]["from"]["row"]
                row_to_image[row_num] = {"base64": img["image_base64"]}
                if "path" in img:
                    row_to_image[row_num]["path"] = img["path"]

        # Sort images by row number
        sorted_rows = sorted(row_to_image.keys())

        # Assign images to data rows where appropriate
        # This is a simple heuristic - we assume images are in order with data rows
        for i, row_data in enumerate(rows):
            if i < len(sorted_rows):
                # Check if there's an image field in the row
                for field, value in row_data.items():
                    # Look for image fields (either by field type or by field name)
                    is_image_field = False

                    # Check field types metadata if available
                    if "_field_types" in row_data and field in row_data["_field_types"]:
                        is_image_field = row_data["_field_types"][field] == "image"

                    # Also check by common image field names
                    if not is_image_field and field.lower() in (
                        "image",
                        "logo",
                        "picture",
                        "photo",
                    ):
                        is_image_field = True

                    if is_image_field:
                        # This is an image field, assign the image data
                        row_data[field] = row_to_image[sorted_rows[i]]
                        break

                # If no explicit image field was found but we have an image, add it to an 'image' field
                if "image" not in row_data:
                    row_data["image"] = row_to_image[sorted_rows[i]]

    def _link_images_to_key_value_pairs(
        self, data: Dict[str, Any], images: List[Dict[str, Any]]
    ) -> None:
        """Link images to key-value pairs based on exact cell position information."""
        if not images or not data:
            return

        # Get field position metadata if available
        field_positions = data.get("_field_positions", {})

        # Process each field that has position information
        for field, position in field_positions.items():
            if field in data and data[field] is None:
                # Look for an image at this exact position
                image_data = self._get_image_at_position(
                    position["row"], position["col"], images
                )
                if image_data:
                    data[field] = image_data
                else:
                    # No image found at the exact position, use text fallback if available
                    # This handles cases where a field is marked as "image" type but contains text
                    text_fallback = position.get("text_fallback")
                    if text_fallback is not None:
                        data[field] = text_fallback
                        logger.debug(
                            f"Using text fallback for image field '{field}': {text_fallback}"
                        )
                    else:
                        logger.debug(
                            f"No image found at position for field '{field}' at row {position['row']}, col {position['col']}"
                        )

        # Fallback: Use the old logic for any remaining image fields without position data
        # This maintains backward compatibility for other parts of the system
        remaining_image_fields = []
        for field, value in data.items():
            if field.startswith("_"):  # Skip metadata fields
                continue

            # Look for image fields that weren't processed above
            is_image_field = False
            if "_field_types" in data and field in data["_field_types"]:
                is_image_field = data["_field_types"][field] == "image"
            elif field.lower() in ("image", "logo", "picture", "photo"):
                is_image_field = True

            if is_image_field and value is None and field not in field_positions:
                remaining_image_fields.append(field)

        # If there are remaining image fields without position data, use the old sequential logic
        if remaining_image_fields:
            # Create a mapping of row numbers to images for remaining fields
            row_to_image = {}
            for img in images:
                if "position" in img and "coordinates" in img["position"]:
                    row_num = img["position"]["coordinates"]["from"]["row"]
                    row_to_image[row_num] = {"base64": img["image_base64"]}
                    if "path" in img:
                        row_to_image[row_num]["path"] = img["path"]

            # Sort images by row number
            sorted_rows = sorted(row_to_image.keys())

            # Assign images to remaining fields sequentially
            image_index = 0
            for field in remaining_image_fields:
                if image_index < len(sorted_rows):
                    data[field] = row_to_image[sorted_rows[image_index]]
                    image_index += 1

    def _get_sheet_metadata(
        self,
        worksheet: Worksheet,
        config: Dict,
        method: str,
        extracted_data: Dict[str, Any],
    ) -> Dict[str, Any]:
        """Get metadata about the processed sheet."""

        types_detected = []
        for subtable in config.get("subtables", []):
            subtable_type = subtable.get("type", "unknown")
            if subtable_type not in types_detected:
                types_detected.append(subtable_type)

        # Count extracted rows
        extracted_rows = 0
        for subtable_name, subtable_data in extracted_data.items():
            if isinstance(subtable_data, list):
                extracted_rows += len(subtable_data)

        return {
            "total_rows": worksheet.max_row,
            "extracted_rows": extracted_rows,
            "total_columns": worksheet.max_column,
            "method": method,
            "types": types_detected,
            "subtables_detected": len(config.get("subtables", [])),
        }

    def cleanup_images(
        self, images: Dict[str, List[Dict[str, Any]]], config: Dict[str, Any] = None
    ) -> None:
        """Clean up extracted images based on configuration.

        Args:
            images: Dictionary of extracted images by sheet name
            config: Configuration dictionary with image storage settings
        """
        if not images:
            return

        # Get app configuration to determine environment
        from .config_manager import ConfigManager

        config_manager = ConfigManager()
        app_config = config_manager.get_app_config()
        development_mode = app_config.get("development_mode", False)

        # Determine if we should clean up images
        if config is None:
            config = {}

        cleanup = False
        if development_mode:
            cleanup = (
                config.get("image_storage", {})
                .get("development_mode", {})
                .get("cleanup_after_merge", False)
            )
        else:
            cleanup = (
                config.get("image_storage", {})
                .get("production_mode", {})
                .get("cleanup_after_merge", True)
            )

        if not cleanup:
            logger.debug("Image cleanup skipped based on configuration")
            return

        # Delete image files
        for sheet_name, sheet_images in images.items():
            for img in sheet_images:
                try:
                    if "path" in img and os.path.exists(img["path"]):
                        os.remove(img["path"])
                        logger.debug(f"Deleted image: {img['path']}")
                except Exception as e:
                    logger.warning(
                        f"Failed to delete image {img.get('path', 'unknown')}: {e}"
                    )

        logger.info("Image cleanup completed")

    def close(self) -> None:
        """Close the workbook and free resources."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None

        # Clean up memory file and image cache
        self._memory_file = None
        self._image_cache = {}

    def _detect_image_format(self, image_data: bytes) -> str:
        """Detect image format from binary data.

        Args:
            image_data: Binary image data

        Returns:
            Image format as a string (e.g., 'png', 'jpg')
        """
        try:
            # Use PIL to detect the image format
            with io.BytesIO(image_data) as image_stream:
                pil_image = PILImage.open(image_stream)
                if pil_image.format:
                    return pil_image.format.lower()
                else:
                    logger.warning("Could not detect image format, defaulting to png")
                    return "png"
        except Exception as e:
            logger.error(f"Error detecting image format: {e}")
            return "png"  # Default to PNG

    def _encode_image_as_base64(self, image_data: bytes, img_format: str) -> str:
        """Encode image data as base64 string with proper MIME type.

        Args:
            image_data: Binary image data
            img_format: Image format (e.g., 'png', 'jpeg')

        Returns:
            Base64 encoded string with data URI prefix
        """
        try:
            # Normalize format for MIME type
            mime_format = img_format.lower()
            if mime_format == "jpg":
                mime_format = "jpeg"

            # Define MIME type mapping
            mime_types = {
                "png": "image/png",
                "jpeg": "image/jpeg",
                "gif": "image/gif",
                "webp": "image/webp",
                "bmp": "image/bmp",
                "tiff": "image/tiff",
            }

            mime_type = mime_types.get(mime_format, "image/png")

            # Encode to base64
            base64_encoded = base64.b64encode(image_data).decode("utf-8")

            # Return data URI
            return f"data:{mime_type};base64,{base64_encoded}"

        except Exception as e:
            logger.error(f"Error encoding image as base64: {e}")
            # Return empty data URI as fallback
            return "data:image/png;base64,"

    def _normalize_image_filename(
        self, sheet_name: str, idx: int, format_name: str
    ) -> str:
        """Normalize image filename for consistency and reliability."""
        # Normalize sheet name: lowercase, replace spaces with underscores, remove special chars
        normalized_sheet = re.sub(r"[^\w_]", "", sheet_name.lower().replace(" ", "_"))
        # Ensure format is lowercase and without leading dot
        format_lower = format_name.lower().lstrip(".")
        # Create normalized filename
        return f"{normalized_sheet}_image_{idx}.{format_lower}"

    def _extract_image_position(self, img) -> Dict[str, Any]:
        """Extract position information from image object."""
        position_info = {
            "anchor_type": None,
            "from_cell": None,
            "to_cell": None,
            "coordinates": None,
            "estimated_cell": None,
            "size_info": None,
        }

        try:
            if hasattr(img, "anchor"):
                anchor = img.anchor
                position_info["anchor_type"] = type(anchor).__name__

                # Two-cell anchor (most common)
                if hasattr(anchor, "_from") and hasattr(anchor, "to"):
                    from_info = anchor._from
                    to_info = anchor.to

                    # Convert to Excel cell references
                    from_col = get_column_letter(from_info.col + 1)
                    from_cell = f"{from_col}{from_info.row + 1}"
                    to_col = get_column_letter(to_info.col + 1)
                    to_cell = f"{to_col}{to_info.row + 1}"

                    position_info.update(
                        {
                            "from_cell": from_cell,
                            "to_cell": to_cell,
                            "coordinates": {
                                "from": {"col": from_info.col, "row": from_info.row},
                                "to": {"col": to_info.col, "row": to_info.row},
                            },
                        }
                    )

                    # Use from_cell as the primary estimated position
                    position_info["estimated_cell"] = from_cell

                    # Calculate span information
                    col_span = to_info.col - from_info.col + 1
                    row_span = to_info.row - from_info.row + 1
                    position_info["size_info"] = {
                        "column_span": col_span,
                        "row_span": row_span,
                    }

                # One-cell anchor
                elif hasattr(anchor, "_from"):
                    from_info = anchor._from
                    from_col = get_column_letter(from_info.col + 1)
                    from_cell = f"{from_col}{from_info.row + 1}"

                    position_info.update(
                        {
                            "from_cell": from_cell,
                            "estimated_cell": from_cell,
                            "coordinates": {
                                "from": {"col": from_info.col, "row": from_info.row}
                            },
                        }
                    )

                    position_info["size_info"] = {"column_span": 1, "row_span": 1}

        except Exception as e:
            logger.debug(f"Could not extract detailed position info: {e}")
            # Fallback to basic position estimation
            position_info["estimated_cell"] = "A1"  # Default fallback
            position_info["size_info"] = {"column_span": 1, "row_span": 1}

        return position_info

    def _auto_detect_sheet_structure(
        self, worksheet: Worksheet, scan_all_rows: bool = False
    ) -> Dict[str, Any]:
        """Auto-detect data structure in worksheet.

        Args:
            worksheet: The worksheet to analyze
            scan_all_rows: If True, scan all rows in the worksheet instead of limiting
        """

        # Scan worksheet to find data patterns
        data_regions = self._scan_data_regions(worksheet, scan_all_rows=scan_all_rows)

        subtables = []
        for region in data_regions:
            subtable_config = self._create_subtable_config(
                worksheet, region, scan_all_rows=scan_all_rows
            )
            if subtable_config:
                subtables.append(subtable_config)

        return {"subtables": subtables}

    def auto_detect_all_sheets(self) -> Dict[str, Any]:
        """Auto-detect data structure for all sheets in the workbook.

        Returns:
            Dictionary with global_settings and sheet_configs for all sheets
        """
        try:
            logger.info("Auto-detecting structure for all sheets in workbook")

            # Create minimal global settings for auto-detection
            global_settings = {
                "image_extraction": {
                    "enabled": True,
                    "formats": ["png", "jpg", "jpeg", "gif", "webp"],
                    "save_format": "png",
                    "cleanup_after_merge": True,
                }
            }

            # Auto-detect each sheet
            sheet_configs = {}
            for sheet_name in self.workbook.sheetnames:
                logger.info(f"Auto-detecting structure for sheet: {sheet_name}")
                worksheet = self.workbook[sheet_name]

                try:
                    # Use existing auto-detection logic with scan_all_rows=True
                    sheet_config = self._auto_detect_sheet_structure(
                        worksheet, scan_all_rows=True
                    )
                    sheet_configs[sheet_name] = sheet_config
                    logger.debug(
                        f"Auto-detected config for {sheet_name}: {sheet_config}"
                    )
                except Exception as e:
                    logger.warning(f"Failed to auto-detect sheet '{sheet_name}': {e}")
                    # Continue with other sheets, don't fail completely
                    continue

            if not sheet_configs:
                raise ExcelProcessingError("Failed to auto-detect any sheet structures")

            logger.info(
                f"Successfully auto-detected structure for {len(sheet_configs)} sheets"
            )

            return {"global_settings": global_settings, "sheet_configs": sheet_configs}

        except Exception as e:
            logger.error(f"Auto-detection failed: {e}")
            raise ExcelProcessingError(f"Failed to auto-detect workbook structure: {e}")

    def _scan_data_regions(
        self,
        worksheet: Worksheet,
        max_scan_rows: int = 100,
        scan_all_rows: bool = False,
    ) -> List[Dict]:
        """Scan worksheet to identify distinct data regions.

        Args:
            worksheet: The worksheet to analyze
            max_scan_rows: Maximum number of rows to scan if scan_all_rows is False
            scan_all_rows: If True, scan all rows in the worksheet
        """
        regions = []
        used_rows = set()  # Track rows already part of a region

        # Determine how many rows to scan
        rows_to_scan = (
            worksheet.max_row
            if scan_all_rows
            else min(max_scan_rows, worksheet.max_row)
        )

        # Simple algorithm: look for header-like patterns
        for row in range(1, rows_to_scan + 1):
            # Skip if this row is already part of another region
            if row in used_rows:
                continue

            # Check if this row looks like headers
            row_cells = [
                worksheet.cell(row=row, column=col).value
                for col in range(1, min(20, worksheet.max_column + 1))
            ]

            # Filter non-empty cells
            non_empty = [cell for cell in row_cells if cell is not None]

            if len(non_empty) >= 2:  # Potential header row
                # Analyze the pattern below this row
                region_info = self._analyze_region_below(
                    worksheet, row, len(non_empty), scan_all_rows=scan_all_rows
                )
                if region_info and region_info["rows"] > 0:
                    regions.append(
                        {
                            "header_row": row,
                            "header_col": 1,
                            "type": region_info["type"],
                            "rows": region_info["rows"],
                            "cols": region_info["cols"],
                        }
                    )

                    # Mark rows as used to avoid overlapping regions
                    for r in range(row, row + region_info["rows"] + 1):
                        used_rows.add(r)

        # Prioritize larger regions and tables over key-value pairs
        regions.sort(
            key=lambda x: (x["type"] == "table", x["rows"] * x["cols"]), reverse=True
        )

        # Return only the best region to avoid confusion
        return regions[:1] if regions else []

    def _analyze_region_below(
        self,
        worksheet: Worksheet,
        header_row: int,
        header_cols: int,
        scan_all_rows: bool = False,
    ) -> Optional[Dict]:
        """Analyze data pattern below a potential header row.

        Args:
            worksheet: The worksheet to analyze
            header_row: Row number of the potential header
            header_cols: Number of columns in the header
            scan_all_rows: If True, scan all rows in the worksheet
        """

        # Look at rows below to determine pattern
        data_rows = 0
        consistent_structure = True
        consecutive_empty_rows = 0

        # Determine max rows to scan
        max_rows_to_scan = (
            worksheet.max_row
            if scan_all_rows
            else min(header_row + 50, worksheet.max_row)
        )

        for row in range(header_row + 1, max_rows_to_scan + 1):
            row_data = [
                worksheet.cell(row=row, column=col).value
                for col in range(1, header_cols + 1)
            ]

            non_empty_count = sum(1 for cell in row_data if cell is not None)

            if non_empty_count == 0:
                consecutive_empty_rows += 1
                # Allow up to 2 consecutive empty rows within a table, break after 2
                if consecutive_empty_rows > 2 and data_rows > 0 and not scan_all_rows:
                    break  # End of data after 2+ consecutive empty rows
            else:
                consecutive_empty_rows = 0  # Reset counter when we find data
                data_rows += 1

                # Check if structure is consistent with table format
                # Lowered threshold to 15% to better handle sparse table data
                if non_empty_count < header_cols * 0.15:  # Less than 15% filled
                    consistent_structure = False

        if data_rows >= 1:
            # Strongly prefer table format for any structured data
            # Only use key_value_pairs for very sparse data (< 10% consistent)
            use_table = consistent_structure or data_rows >= 2
            detection_type = "table" if use_table else "key_value_pairs"
            logger.debug(
                f"Detection result: data_rows={data_rows}, consistent={consistent_structure}, use_table={use_table}, type={detection_type}"
            )
            return {
                "type": detection_type,
                "rows": data_rows,
                "cols": header_cols,
            }

        return None

    def _create_subtable_config(
        self, worksheet: Worksheet, region: Dict, scan_all_rows: bool = False
    ) -> Optional[Dict]:
        """Create subtable configuration based on detected region.

        Args:
            worksheet: The worksheet to analyze
            region: Region information from _scan_data_regions
            scan_all_rows: If True, configure to extract all rows
        """

        header_row = region["header_row"]
        header_col = region["header_col"]

        # Get headers
        headers = []
        for col in range(header_col, header_col + region["cols"]):
            cell_value = worksheet.cell(row=header_row, column=col).value
            if cell_value:
                headers.append(str(cell_value).strip())

        if not headers:
            return None

        # Create column mappings
        column_mappings = {}
        for header in headers:
            column_mappings[header] = normalize_column_name(header)

        # Determine search text for this region
        search_text = headers[0] if headers else "Data"

        # Determine max_rows based on scan_all_rows flag
        max_rows = worksheet.max_row if scan_all_rows else region["rows"] + 10

        config = {
            "name": f"auto_detected_{region['type']}_{header_row}",
            "type": region["type"],
            "header_search": {
                "method": "contains_text",
                "text": search_text,
                "column": "A",
                "search_range": f"A{max(1, header_row-2)}:A{header_row+2}",
            },
            "data_extraction": {
                "headers_row_offset": 0,
                "data_row_offset": 1,
                "max_columns": region["cols"],
                "max_rows": max_rows,
                "column_mappings": column_mappings,
            },
        }

        # Add orientation for key-value pairs
        if region["type"] == "key_value_pairs":
            config["data_extraction"]["orientation"] = "horizontal"

        return config

    def _initialize_range_exporter(self, sharepoint_config: Dict[str, Any]) -> None:
        """Initialize range exporter for local processing (no longer requires Graph API)."""
        if self._range_exporter is None:
            range_image_logger.info("🔧 Initializing local ExcelRangeExporter")

            self._range_exporter = ExcelRangeExporter()

            # Set debug directory for development mode if available
            debug_dir = getattr(self, "_debug_directory", None)
            if debug_dir:
                self._range_exporter.set_debug_directory(debug_dir)
                range_image_logger.info(
                    f"🗂️ Set debug directory for range images: {debug_dir}"
                )

            range_image_logger.info("✅ ExcelRangeExporter initialized successfully")

    def _extract_range_images(
        self,
        range_configs_data: List[Dict[str, Any]],
        sharepoint_config: Dict[str, Any],
    ) -> Dict[str, str]:
        """Extract range images using Graph API.

        Args:
            range_configs_data: List of range image configurations
            sharepoint_config: SharePoint configuration from global settings

        Returns:
            Dictionary mapping field_name to image file path
        """
        # CRITICAL DEBUG: Log entry point and workbook state
        range_image_logger.info("=" * 80)
        range_image_logger.info("🚀 _extract_range_images ENTRY POINT")
        range_image_logger.info(f"📊 Configs: {len(range_configs_data) if range_configs_data else 0}")
        range_image_logger.info(f"📝 SharePoint config: {sharepoint_config}")
        range_image_logger.info(f"🔍 DEBUG: _is_memory_file = {getattr(self, '_is_memory_file', 'NOT_SET')}")
        range_image_logger.info(f"🔍 DEBUG: _memory_file exists = {hasattr(self, '_memory_file') and self._memory_file is not None}")
        range_image_logger.info(f"🔍 DEBUG: workbook exists = {hasattr(self, 'workbook') and self.workbook is not None}")
        range_image_logger.info("=" * 80)

        # Check if SharePoint is properly configured
        # For URL-based approach, we can proceed even if sharepoint.enabled is false
        using_url_based_approach = (
            hasattr(self, "_url_file_source") and self._url_file_source
        )

        if not sharepoint_config.get("enabled") and not using_url_based_approach:
            range_image_logger.warning(
                "⚠️ SharePoint not enabled in configuration and not using URL-based approach - range image extraction skipped"
            )
            range_image_logger.info(
                "💡 For range extraction, either enable SharePoint config or use SharePoint URLs"
            )
            return {}

        range_image_logger.info(
            f"📝 SharePoint site_id in config: {sharepoint_config.get('site_id')}"
        )
        range_image_logger.info(
            f"📝 Using URL-based approach: {using_url_based_approach}"
        )

        if not sharepoint_config.get("site_id") and not using_url_based_approach:
            range_image_logger.error(
                "❌ SharePoint site_id not configured and not using URL-based approach - range image extraction failed"
            )
            range_image_logger.info(
                "💡 NOTE: When using SharePoint URLs, you can either:"
            )
            range_image_logger.info(
                "   1. Configure site_id and drive_id in config.json"
            )
            range_image_logger.info(
                "   2. Use SharePoint Doc.aspx URLs (URL-based approach)"
            )
            return {}

        # Initialize range exporter with SharePoint configuration
        # For URL-based approach, we need modified SharePoint config
        if using_url_based_approach:
            # Create a modified sharepoint config that doesn't require site_id for initialization
            modified_sharepoint_config = sharepoint_config.copy()
            modified_sharepoint_config["enabled"] = True
            # Remove site_id requirement for URL-based operations
            range_image_logger.info(
                "🔧 Modifying SharePoint config for URL-based range extraction"
            )
            self._initialize_range_exporter(modified_sharepoint_config)
        else:
            self._initialize_range_exporter(sharepoint_config)

        if not self._range_exporter:
            range_image_logger.error(
                "❌ Range exporter not initialized - Graph API credentials required"
            )
            return {}

        if not range_configs_data:
            range_image_logger.warning("⚠️ No range configurations provided")
            return {}

        try:
            # Log each range configuration with detailed formatting
            for i, config_data in enumerate(range_configs_data):
                log_range_config(config_data, i)

            # Validate and create range configurations
            range_configs = create_range_configs_from_dict(range_configs_data)
            range_image_logger.info(
                f"✅ CONFIGURATION VALIDATION PASSED - Processing {len(range_configs)} range configurations"
            )

            # Log available sheet names for debugging
            available_sheets = list(self.workbook.sheetnames) if self.workbook else []
            range_image_logger.info(f"📊 AVAILABLE SHEETS: {available_sheets}")

            # Log requested sheet names from configs
            requested_sheets = [config.sheet_name for config in range_configs]
            range_image_logger.info(f"🎯 REQUESTED SHEETS: {requested_sheets}")

            # Export ranges as images
            range_image_logger.info(
                f"🚀 STARTING EXPORT PROCESS with {len(range_configs)} configurations"
            )

            if using_url_based_approach:
                # Use URL-based approach for SharePoint files
                range_image_logger.info(
                    f"🌐 Using URL-based range extraction for SharePoint file"
                )
                range_image_logger.info("🔥 ABOUT TO CALL _export_ranges_from_sharepoint_url")
                try:
                    results = self._export_ranges_from_sharepoint_url(range_configs)
                    range_image_logger.info("✅ _export_ranges_from_sharepoint_url COMPLETED SUCCESSFULLY")
                except Exception as e:
                    range_image_logger.error(f"💥 _export_ranges_from_sharepoint_url FAILED: {e}")
                    range_image_logger.error(f"💥 Exception type: {type(e).__name__}")
                    import traceback
                    range_image_logger.error(f"💥 Traceback: {traceback.format_exc()}")
                    raise
            else:
                # Use traditional file upload approach
                excel_file_path = self._get_excel_file_path()
                range_image_logger.info(f"📁 EXCEL FILE PATH: {excel_file_path}")
                if not excel_file_path:
                    range_image_logger.error(
                        "❌ Cannot determine Excel file path for range image export"
                    )
                    return {}
                results = self._range_exporter.export_ranges(
                    excel_file_path, range_configs
                )

            # Process results and return mapping
            range_images = {}
            successful_exports = 0
            total_exports = len(results)

            for i, result in enumerate(results, 1):
                if result.success:
                    range_images[result.field_name] = result.image_path
                    successful_exports += 1
                    log_range_export_progress(
                        i, total_exports, result.field_name, "success"
                    )
                    range_image_logger.info(
                        f"✅ EXPORT SUCCESS: {result.field_name}\n"
                        f"   📁 Path: {result.image_path}\n"
                        f"   📐 Dimensions: {result.width} x {result.height}\n"
                        f"   📊 Range Size: {result.range_dimensions[0]} rows x {result.range_dimensions[1]} cols"
                    )
                else:
                    log_range_export_progress(
                        i, total_exports, result.field_name, "failed"
                    )
                    range_image_logger.error(
                        f"❌ EXPORT FAILED: {result.field_name}\n"
                        f"   💥 Error: {result.error_message}"
                    )

            range_image_logger.info(
                f"🏁 EXTRACTION COMPLETED\n"
                f"   ✅ Successful: {successful_exports}/{total_exports}\n"
                f"   📁 Output Files: {list(range_images.keys())}"
            )
            return range_images

        except Exception as e:
            range_image_logger.error(
                f"💥 CRITICAL ERROR during range image extraction: {e}"
            )
            return {}

    def _get_excel_file_path(self) -> Optional[str]:
        """Get the Excel file path for range export operations."""
        if self._is_memory_file:
            # For in-memory files, we need to save to a temporary file
            if self._memory_file:
                try:
                    import tempfile

                    with tempfile.NamedTemporaryFile(
                        delete=False, suffix=".xlsx"
                    ) as tmp_file:
                        tmp_file.write(self._memory_file.getvalue())
                        temp_path = tmp_file.name
                    logger.info(f"Created temporary file for range export: {temp_path}")
                    return temp_path
                except Exception as e:
                    logger.error(
                        f"Failed to create temporary file for range export: {e}"
                    )
                    return None
            return None
        else:
            return self.file_path

    def cleanup_range_exporter(self) -> None:
        """Cleanup resources used by range exporter."""
        if self._range_exporter:
            try:
                self._range_exporter.cleanup_temp_files()
                logger.info("Cleaned up range exporter temporary files")
            except Exception as e:
                logger.warning(f"Failed to cleanup range exporter: {e}")

    def _export_ranges_from_sharepoint_url(self, range_configs: List) -> List:
        """Export range images from SharePoint URL using downloaded file.

        Args:
            range_configs: List of RangeImageConfig objects

        Returns:
            List of RangeImageResult objects
        """
        from .excel_range_exporter import RangeImageResult

        # CRITICAL DEBUG: Method entry logging
        logger.info("🚀 _export_ranges_from_sharepoint_url METHOD CALLED")
        range_image_logger.info(
            f"🌐 Starting URL-based range extraction from SharePoint"
        )

        try:
            # Since we're using local processing now, we'll use the workbook already loaded
            # The workbook should already be downloaded and loaded in self.workbook
            if not self.workbook:
                range_image_logger.error(
                    "❌ No workbook loaded for range extraction"
                )
                return []

            # Create a temporary file from the workbook for the range exporter
            import tempfile
            import os
            
            temp_file = None
            try:
                # Create temporary file for range extraction
                temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
                
                # Debug: Check memory file conditions
                range_image_logger.info(f"🔍 DEBUG: _is_memory_file={getattr(self, '_is_memory_file', 'NOT_SET')}")
                range_image_logger.info(f"🔍 DEBUG: _memory_file exists={hasattr(self, '_memory_file') and self._memory_file is not None}")
                if hasattr(self, '_memory_file') and self._memory_file is not None:
                    range_image_logger.info(f"🔍 DEBUG: _memory_file size={len(self._memory_file.getvalue())} bytes")
                
                # Check if workbook was loaded from memory
                if self._is_memory_file and self._memory_file:
                    # Copy original memory file directly instead of saving workbook
                    # This avoids the "I/O operation on closed file" error
                    range_image_logger.info(f"📁 Using MEMORY FILE path - copying original data")
                    memory_data = self._memory_file.getvalue()
                    temp_file.write(memory_data)
                    temp_file.close()
                    range_image_logger.info(f"📁 Copied {len(memory_data)} bytes to temp file: {temp_file.name}")
                else:
                    # For file-based workbooks, save normally
                    range_image_logger.info(f"📁 Using FILE-BASED path - calling workbook.save()")
                    temp_file.close()  # Close before saving
                    range_image_logger.info(f"📁 About to save workbook to: {temp_file.name}")
                    self.workbook.save(temp_file.name)
                    range_image_logger.info(f"📁 Successfully saved workbook to temp file")
                
                # Use the standard export_ranges method
                range_image_logger.info(f"📊 About to call export_ranges with {len(range_configs)} configs")
                range_image_logger.info(f"📊 Temp file path: {temp_file.name}")
                results = self._range_exporter.export_ranges(temp_file.name, range_configs)
                range_image_logger.info(f"📊 export_ranges completed successfully with {len(results)} results")
                
                return results
                
            finally:
                # Clean up temp file
                if temp_file and os.path.exists(temp_file.name):
                    try:
                        os.unlink(temp_file.name)
                    except:
                        pass
                        
        except Exception as e:
            range_image_logger.error(f"❌ URL-based range extraction failed [SOURCE:_export_ranges_from_sharepoint_url]: {e}")
            # Return failed results for all configs
            from .excel_range_exporter import RangeImageResult

            return [
                RangeImageResult(
                    field_name=config.field_name,
                    image_path="",
                    image_data=b"",
                    width=0,
                    height=0,
                    range_dimensions=(0, 0),
                    success=False,
                    error_message=str(e),
                )
                for config in range_configs
            ]
