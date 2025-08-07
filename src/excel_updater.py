"""Excel file update module for modifying cells with new data."""

import logging
import os
import io
import uuid
import re
import base64
import copy
import requests
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple, Union
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
from PIL import Image as PILImage

from .temp_file_manager import TempFileManager
from .utils.exceptions import ExcelProcessingError, ValidationError

logger = logging.getLogger(__name__)


class ExcelUpdater:
    """Updates Excel files with new data according to configuration."""

    def __init__(self, excel_file_path: str):
        """Initialize Excel updater with file path."""
        self.file_path = excel_file_path
        # Load workbook with minimal parameters to avoid corruption
        # Use default settings that are most compatible
        self.workbook = load_workbook(excel_file_path)
        self.update_log = []
        self.temp_image_files = []  # Track temp files for cleanup

    def update_excel(
        self,
        update_data: Dict[str, Any],
        config: Dict[str, Any],
        include_update_log: bool = False,
    ) -> str:
        """
        Main update method - returns path to updated file.

        Args:
            update_data: Dictionary with data to update
            config: Configuration mapping for updates
            include_update_log: Whether to include diagnostic update_log sheet (default: False)

        Returns:
            Path to updated Excel file
        """
        try:
            self._log_info("Starting Excel update process")

            # Check existing content for preservation
            self._log_existing_content()

            # Validate configuration
            self._validate_update_config(config)

            # Get list of sheets that will be updated
            sheets_to_update = set(config.get("sheet_configs", {}).keys())
            self._log_info(f"Sheets to be updated: {sheets_to_update}")

            # Process each sheet using two-phase approach
            for sheet_name, sheet_config in config.get("sheet_configs", {}).items():
                if sheet_name not in self.workbook.sheetnames:
                    self._log_error(f"Sheet '{sheet_name}' not found in workbook")
                    continue

                sheet = self.workbook[sheet_name]
                self._log_info(f"Processing sheet: {sheet_name}")

                # Phase 1: Detect all subtables before making any changes
                detected_subtables = self._detect_all_subtables(sheet, sheet_config)
                
                # Phase 2: Process subtables in correct order with position tracking
                self._process_subtables_in_order(sheet, detected_subtables, update_data)

            # Verify non-updated sheets remain intact
            for sheet_name in self.workbook.sheetnames:
                if sheet_name not in sheets_to_update and sheet_name != "update_log":
                    self._verify_sheet_preservation(sheet_name)

            # Add diagnostic log sheet if requested
            if include_update_log:
                self._add_update_log_sheet()

            # Save updated workbook
            output_path = self._save_updated_workbook()
            self._log_info(f"Excel update completed successfully: {output_path}")

            return output_path

        except Exception as e:
            self._log_error(f"Update process failed: {e}")
            raise ExcelProcessingError(f"Failed to update Excel file: {e}")

    def _update_subtable(
        self,
        sheet: Worksheet,
        subtable_config: Dict[str, Any],
        update_data: Dict[str, Any],
    ) -> None:
        """Update subtable with unified offset handling."""
        subtable_name = subtable_config["name"]
        subtable_type = subtable_config["type"]

        self._log_info(f"Updating subtable: {subtable_name}")

        # Skip if no data_update configuration
        if "data_update" not in subtable_config:
            self._log_warning(
                f"Skipping subtable '{subtable_name}' - no data_update configuration"
            )
            return

        # Find starting location
        location = self._find_update_location(sheet, subtable_config["header_search"])
        if not location["found"]:
            self._log_error(f"Could not find location for subtable '{subtable_name}'")
            return

        # Get data for this subtable
        if subtable_name not in update_data:
            self._log_warning(f"No data provided for subtable '{subtable_name}'")
            return

        data = update_data[subtable_name]
        update_config = subtable_config["data_update"]

        # Apply offsets to starting location
        base_row = location["row"] + update_config.get("headers_row_offset", 0)
        base_col = location["col"] + update_config.get("headers_col_offset", 0)

        # Update based on subtable type
        if subtable_type == "key_value_pairs":
            self._update_key_value_pairs_with_offsets(
                sheet, base_row, base_col, update_config, data
            )
        elif subtable_type == "table":
            self._update_table_with_offsets(
                sheet, base_row, base_col, update_config, data
            )
        elif subtable_type == "matrix_table":
            self._update_matrix_table_with_offsets(
                sheet, base_row, base_col, update_config, data
            )

    def _find_update_location(
        self, sheet: Worksheet, header_search_config: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Find starting location using either method."""
        method = header_search_config["method"]

        if method == "cell_address":
            return self._find_by_cell_address(sheet, header_search_config)
        elif method == "contains_text":
            return self._find_by_contains_text(sheet, header_search_config)
        else:
            raise ValueError(f"Unsupported header search method: {method}")

    def _find_by_cell_address(
        self, sheet: Worksheet, config: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Direct cell address method."""
        cell_address = config["cell"]
        try:
            row, col = self._parse_cell_address(cell_address)
            self._log_info(
                f"Found location by cell address {cell_address}: row={row}, col={col}"
            )
            return {"row": row, "col": col, "found": True, "address": cell_address}
        except Exception as e:
            self._log_error(f"Invalid cell address {cell_address}: {e}")
            return {"found": False}

    def _find_by_contains_text(
        self, sheet: Worksheet, config: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Text search method."""
        search_text = config["text"]
        search_column = config["column"]
        search_range = config.get(
            "search_range", f"{search_column}1:{search_column}100"
        )

        try:
            for cell in sheet[search_range]:
                for cell_obj in cell:
                    if (
                        cell_obj.value
                        and search_text.lower() in str(cell_obj.value).lower()
                    ):
                        self._log_info(
                            f"Found location by text '{search_text}' at {cell_obj.coordinate}"
                        )
                        return {
                            "row": cell_obj.row,
                            "col": cell_obj.column,
                            "found": True,
                            "address": cell_obj.coordinate,
                        }

            self._log_error(f"Text '{search_text}' not found in range {search_range}")
            return {"found": False}

        except Exception as e:
            self._log_error(f"Text search failed: {e}")
            return {"found": False}

    def _update_key_value_pairs_with_offsets(
        self,
        sheet: Worksheet,
        base_row: int,
        base_col: int,
        config: Dict[str, Any],
        data: Dict[str, Any],
    ) -> None:
        """Update key-value pairs with offset support."""
        column_mappings = config["column_mappings"]
        orientation = config.get("orientation", "horizontal")
        data_row_offset = config.get("data_row_offset", 1)
        data_col_offset = config.get("data_col_offset", 1)

        if orientation == "horizontal":
            # Headers are in base_row, data in base_row + data_row_offset
            data_row = base_row + data_row_offset

            for mapping_key, field_config in column_mappings.items():
                field_name = field_config["name"]
                field_type = field_config["type"]

                if field_name in data:
                    # Handle both cell_address and column-based mappings
                    if self._is_cell_address(mapping_key):
                        # Direct cell address (e.g., "B14")
                        cell_address = mapping_key
                    else:
                        # Column header name - find the column
                        header_col = self._find_header_column(
                            sheet, base_row, mapping_key, config, base_col
                        )
                        if header_col:
                            cell_address = f"{get_column_letter(header_col)}{data_row}"
                        else:
                            self._log_error(f"Header '{mapping_key}' not found")
                            continue

                    self._update_cell(sheet, cell_address, data[field_name], field_type)

        elif orientation == "vertical":
            # Keys are in base_col, data in base_col + data_col_offset
            data_col = base_col + data_col_offset

            for mapping_key, field_config in column_mappings.items():
                field_name = field_config["name"]
                field_type = field_config["type"]

                if field_name in data:
                    # Find the row where this key is located
                    key_row = self._find_key_row(
                        sheet, base_row, base_col, mapping_key, config
                    )
                    if key_row:
                        cell_address = f"{get_column_letter(data_col)}{key_row}"
                        self._update_cell(
                            sheet, cell_address, data[field_name], field_type
                        )
                    else:
                        self._log_error(
                            f"Key '{mapping_key}' not found in vertical layout"
                        )

    def _update_table_with_offsets(
        self,
        sheet: Worksheet,
        base_row: int,
        base_col: int,
        config: Dict[str, Any],
        data: List[Dict[str, Any]],
    ) -> None:
        """Update table data with offset support and optional style copying."""
        column_mappings = config["column_mappings"]
        orientation = config.get("orientation", "vertical")
        data_row_offset = config.get("data_row_offset", 1)
        copy_first_row_style = config.get("copy_first_row_style", True)

        if orientation == "vertical":
            # Start data from base_row + data_row_offset
            data_start_row = base_row + data_row_offset
            
            # First, determine all column positions we'll be updating
            column_positions = {}
            for mapping_key in column_mappings.keys():
                header_col = self._find_header_column(
                    sheet, base_row, mapping_key, config, base_col
                )
                if header_col:
                    column_positions[mapping_key] = header_col
                elif self._is_column_letter(mapping_key):
                    column_positions[mapping_key] = column_index_from_string(mapping_key)
            
            # Capture styles from the first data row if it exists and option is enabled
            captured_styles = None
            if copy_first_row_style and data_start_row <= sheet.max_row:
                columns_to_capture = list(column_positions.values())
                if columns_to_capture:
                    captured_styles = self._capture_row_styles(
                        sheet, data_start_row, columns_to_capture
                    )
                    self._log_info(
                        f"Captured styles from row {data_start_row} for {len(columns_to_capture)} columns"
                    )

            for row_idx, row_data in enumerate(data):
                current_row = data_start_row + row_idx

                for mapping_key, field_config in column_mappings.items():
                    field_name = field_config["name"]
                    field_type = field_config["type"]

                    if field_name in row_data and mapping_key in column_positions:
                        col_idx = column_positions[mapping_key]
                        cell_address = f"{get_column_letter(col_idx)}{current_row}"
                        
                        # Update the cell value
                        self._update_cell(
                            sheet, cell_address, row_data[field_name], field_type
                        )
                        
                        # Apply captured style if available (skip first row as it already has the style)
                        if captured_styles and row_idx > 0 and col_idx in captured_styles:
                            cell = sheet[cell_address]
                            self._apply_cell_style(cell, captured_styles[col_idx])
                
                # Apply row height if captured
                if captured_styles and row_idx > 0 and captured_styles.get("row_height"):
                    sheet.row_dimensions[current_row].height = captured_styles["row_height"]
                
                # Apply merged cells pattern if captured
                if captured_styles and row_idx > 0 and captured_styles.get("merged_cells"):
                    for merge_info in captured_styles["merged_cells"]:
                        merge_range = f"{get_column_letter(merge_info['min_col'])}{current_row}:" \
                                    f"{get_column_letter(merge_info['max_col'])}{current_row}"
                        try:
                            sheet.merge_cells(merge_range)
                        except:
                            pass  # Skip if merge fails

    def _update_matrix_table_with_offsets(
        self,
        sheet: Worksheet,
        base_row: int,
        base_col: int,
        config: Dict[str, Any],
        data: Dict[str, Dict[str, Any]],
    ) -> None:
        """Update matrix table data with offset support.

        Expects data in format: {row_key: {col_key: value}}
        """
        column_mappings = config["column_mappings"]
        row_key_mappings = config.get("row_key_mappings", {})

        headers_row = base_row + config.get("headers_row_offset", 0)
        data_start_row = headers_row + config.get("data_row_offset", 1)

        # Support for column offset - allows table to start in different column than search text
        headers_col_offset = config.get("headers_col_offset", 0)
        header_col = base_col + headers_col_offset

        # Row keys are in the first column of the data area
        row_keys_col_offset = config.get("row_keys_col_offset", 0)
        row_keys_col = header_col + row_keys_col_offset

        # Data starts in the column after row keys
        data_col_offset = config.get("data_col_offset", 1)
        data_start_col = row_keys_col + data_col_offset

        max_rows = config.get("max_rows", 1000)

        # Create reverse mapping for row keys (JSON key -> Excel key)
        reverse_row_key_mappings = {}
        for excel_key, json_key in row_key_mappings.items():
            reverse_row_key_mappings[json_key] = excel_key

        # Create reverse mapping for column headers (JSON key -> Excel key)
        reverse_column_mappings = {}
        for excel_key, mapping in column_mappings.items():
            if isinstance(mapping, str):
                json_key = mapping
            else:
                json_key = mapping.get("name")
            if json_key:
                reverse_column_mappings[json_key] = excel_key

        # Update data for each row
        for json_row_key, row_data in data.items():
            if json_row_key == "_field_types":  # Skip metadata
                continue

            # Find the Excel row for this row key
            target_row = None

            # Look for the row key in the Excel sheet
            for row_offset in range(max_rows):
                row = data_start_row + row_offset
                row_key_cell = sheet.cell(row=row, column=row_keys_col)

                if row_key_cell.value:
                    excel_row_key = str(row_key_cell.value).strip()

                    # Check if this matches our target row key (either directly or through mapping)
                    if (
                        excel_row_key == json_row_key
                        or reverse_row_key_mappings.get(json_row_key) == excel_row_key
                    ):
                        target_row = row
                        break

            if target_row is None:
                self._log_error(f"Row key '{json_row_key}' not found in matrix table")
                continue

            # Update cells in this row
            for json_col_key, value in row_data.items():
                # Find the Excel column for this column key
                target_col = None

                # Get the column header mapping
                excel_col_key = reverse_column_mappings.get(json_col_key, json_col_key)

                # Find the column in the header row
                header_col_found = self._find_header_column(
                    sheet, headers_row, excel_col_key, config, data_start_col
                )

                if header_col_found:
                    target_col = header_col_found
                else:
                    self._log_error(
                        f"Column '{json_col_key}' (Excel: '{excel_col_key}') not found in matrix table"
                    )
                    continue

                # Get field type from column mappings
                if excel_col_key in column_mappings:
                    mapping = column_mappings[excel_col_key]
                    if isinstance(mapping, str):
                        field_type = "text"
                    else:
                        field_type = mapping.get("type", "text")
                else:
                    field_type = "text"

                # Update the cell
                cell_address = f"{get_column_letter(target_col)}{target_row}"
                self._update_cell(sheet, cell_address, value, field_type)

    def _capture_row_styles(
        self, sheet: Worksheet, row: int, columns: List[int]
    ) -> Dict[int, Dict[str, Any]]:
        """
        Capture all formatting styles from specified columns in a row.
        
        Args:
            sheet: Excel worksheet
            row: Row number to capture styles from
            columns: List of column indices to capture
            
        Returns:
            Dictionary mapping column index to style properties
        """
        styles = {}
        
        for col in columns:
            cell = sheet.cell(row=row, column=col)
            
            # Capture all style properties
            style_info = {
                "font": copy.copy(cell.font) if cell.font else None,
                "fill": copy.copy(cell.fill) if cell.fill else None,
                "border": copy.copy(cell.border) if cell.border else None,
                "alignment": copy.copy(cell.alignment) if cell.alignment else None,
                "number_format": cell.number_format if cell.number_format else None,
            }
            
            styles[col] = style_info
            
        # Also capture row height
        styles["row_height"] = sheet.row_dimensions[row].height
        
        # Capture merged cell information for this row
        merged_ranges = []
        for merged_range in sheet.merged_cells.ranges:
            if merged_range.min_row == row and merged_range.max_row == row:
                merged_ranges.append({
                    "min_col": merged_range.min_col,
                    "max_col": merged_range.max_col
                })
        styles["merged_cells"] = merged_ranges
        
        return styles
    
    def _apply_cell_style(
        self, cell, style_info: Dict[str, Any]
    ) -> None:
        """
        Apply captured style to a cell.
        
        Args:
            cell: The cell to apply styles to
            style_info: Dictionary containing style properties
        """
        if style_info.get("font"):
            cell.font = style_info["font"]
        
        if style_info.get("fill"):
            cell.fill = style_info["fill"]
        
        if style_info.get("border"):
            cell.border = style_info["border"]
        
        if style_info.get("alignment"):
            cell.alignment = style_info["alignment"]
        
        if style_info.get("number_format"):
            cell.number_format = style_info["number_format"]

    def _update_cell(
        self, sheet: Worksheet, cell_address: str, value: Any, field_type: str
    ) -> bool:
        """Update single cell with type-specific handling."""
        try:
            cell = sheet[cell_address]
            original_value = cell.value

            if field_type == "text":
                cell.value = str(value) if value is not None else ""
                self._log_success(f"Updated {cell_address} with text: '{value}'")

            elif field_type == "number":
                try:
                    cell.value = float(value) if value is not None else 0
                    self._log_success(f"Updated {cell_address} with number: {value}")
                except (ValueError, TypeError):
                    cell.value = str(value) if value is not None else ""
                    self._log_warning(
                        f"Could not convert '{value}' to number, saved as text in {cell_address}"
                    )

            elif field_type == "image":
                return self._insert_image_v2(sheet, cell_address, value)

            elif field_type == "formula":
                # Set formula directly - Excel will evaluate it
                cell.value = str(value) if value is not None else ""
                self._log_success(f"Updated {cell_address} with formula: '{value}'")

            else:
                # Default to text
                cell.value = str(value) if value is not None else ""
                self._log_success(
                    f"Updated {cell_address} with default text: '{value}'"
                )

            return True

        except Exception as e:
            self._log_error(f"Failed to update {cell_address}: {e}")
            try:
                # Try to set error value if valid cell address
                if self._is_cell_address(cell_address):
                    sheet[cell_address].value = "!#VALUE"
            except:
                pass  # If even error setting fails, just continue
            return False

    def _insert_image(
        self, sheet: Worksheet, cell_address: str, image_data: Any
    ) -> bool:
        """Insert image above specified cell."""
        temp_image_path = None
        try:
            if not image_data:
                sheet[cell_address].value = "[NO_IMAGE_DATA]"
                self._log_warning(f"No image data provided for {cell_address}")
                return True

            # Handle different image input types
            final_image_path = None
            final_size = None

            if isinstance(image_data, str):
                if image_data.startswith("file://") or os.path.exists(image_data):
                    # Local file path - try to use directly without processing
                    try:
                        file_path = image_data
                        if file_path.startswith("file://"):
                            import urllib.parse

                            file_path = urllib.parse.unquote(file_path[7:])

                        if os.path.exists(file_path):
                            # Load and process for sizing only, but use original file for Excel
                            with open(file_path, "rb") as f:
                                image_bytes = f.read()
                            pil_img = PILImage.open(io.BytesIO(image_bytes))
                            processed_img, final_size = self._process_image_with_sizing(
                                pil_img
                            )

                            # If image needs resizing, create temp file
                            if final_size != pil_img.size:
                                import tempfile

                                temp_fd, temp_image_path = tempfile.mkstemp(
                                    suffix=".png"
                                )
                                os.close(temp_fd)
                                processed_img.save(temp_image_path, format="PNG")
                                self.temp_image_files.append(temp_image_path)
                                final_image_path = temp_image_path
                            else:
                                # Use original file directly
                                final_image_path = file_path
                        else:
                            raise ValueError(f"File not found: {file_path}")
                    except Exception as e:
                        self._log_error(
                            f"Failed to process local file {image_data}: {e}"
                        )
                        sheet[cell_address].value = "!#VALUE"
                        return False

                elif image_data.startswith("data:image"):
                    # Base64 image - decode and process
                    image_bytes = self._decode_base64_image(image_data)
                    pil_img = PILImage.open(io.BytesIO(image_bytes))
                    processed_img, final_size = self._process_image_with_sizing(pil_img)

                    import tempfile

                    temp_fd, temp_image_path = tempfile.mkstemp(suffix=".png")
                    os.close(temp_fd)
                    processed_img.save(temp_image_path, format="PNG")
                    self.temp_image_files.append(temp_image_path)
                    final_image_path = temp_image_path

                elif image_data.startswith("http"):
                    # Download URL and process
                    image_bytes = self._download_image(image_data)
                    pil_img = PILImage.open(io.BytesIO(image_bytes))
                    processed_img, final_size = self._process_image_with_sizing(pil_img)

                    import tempfile

                    temp_fd, temp_image_path = tempfile.mkstemp(suffix=".png")
                    os.close(temp_fd)
                    processed_img.save(temp_image_path, format="PNG")
                    self.temp_image_files.append(temp_image_path)
                    final_image_path = temp_image_path
                else:
                    # Text fallback
                    sheet[cell_address].value = str(image_data)
                    self._log_info(
                        f"Inserted text instead of image at {cell_address}: {image_data}"
                    )
                    return True
            else:
                # Unknown format - use text
                sheet[cell_address].value = str(image_data)
                self._log_warning(
                    f"Unknown image format, inserted as text at {cell_address}"
                )
                return True

            if not final_image_path or not final_size:
                sheet[cell_address].value = "!#VALUE"
                self._log_error(f"Could not process image data for {cell_address}")
                return False

            # Create Excel image from final path
            excel_img = ExcelImage(final_image_path)

            # Position at target cell
            row, col = self._parse_cell_address(cell_address)
            excel_img.anchor = f"{get_column_letter(col)}{row}"

            sheet.add_image(excel_img)
            # Keep cell empty - no text pollution

            # Adjust row height to fit image
            self._adjust_row_height(sheet, row, final_size[1])

            self._log_success(
                f"Inserted image at {cell_address} (size: {final_size[0]}x{final_size[1]})"
            )
            return True

        except Exception as e:
            self._log_error(f"Image insert failed at {cell_address}: {e}")
            try:
                # Try to set error value if valid cell address
                if self._is_cell_address(cell_address):
                    sheet[cell_address].value = "!#VALUE"
            except:
                pass  # If even error setting fails, just continue
            return False

    def _decode_base64_image(self, base64_data: str) -> bytes:
        """Decode base64 image data."""
        try:
            # Extract base64 data after comma (remove data:image/png;base64, prefix)
            if "," in base64_data:
                base64_data = base64_data.split(",", 1)[1]

            return base64.b64decode(base64_data)
        except Exception as e:
            raise ValueError(f"Invalid base64 image data: {e}")

    def _download_image(self, url: str, timeout: int = 10) -> bytes:
        """Download image from URL with timeout."""
        try:
            response = requests.get(url, timeout=timeout, stream=True)
            response.raise_for_status()

            # Check content type
            content_type = response.headers.get("content-type", "")
            if not content_type.startswith("image/"):
                raise ValueError(
                    f"URL does not point to an image (content-type: {content_type})"
                )

            return response.content

        except Exception as e:
            raise ValueError(f"Failed to download image from {url}: {e}")

    def _load_local_image(self, file_path: str) -> bytes:
        """Load image from local file path."""
        try:
            # Handle file:// URLs
            if file_path.startswith("file://"):
                import urllib.parse

                file_path = urllib.parse.unquote(
                    file_path[7:]
                )  # Remove 'file://' prefix

            # Check if file exists
            if not os.path.exists(file_path):
                raise ValueError(f"Image file not found: {file_path}")

            # Check if it's a valid image file
            valid_extensions = {
                ".png",
                ".jpg",
                ".jpeg",
                ".gif",
                ".bmp",
                ".tiff",
                ".webp",
            }
            file_ext = os.path.splitext(file_path.lower())[1]
            if file_ext not in valid_extensions:
                raise ValueError(f"Unsupported image format: {file_ext}")

            # Read file
            with open(file_path, "rb") as f:
                return f.read()

        except Exception as e:
            raise ValueError(f"Failed to load local image from {file_path}: {e}")

    def _process_image_with_sizing(self, pil_img):
        """Process image with size management and return (processed_image, (width, height))."""
        try:
            # Default image size in points (Excel units)
            default_width = 50  # points
            default_height = 25  # points

            # Get original size
            original_width, original_height = pil_img.size

            # Calculate if we need to resize (if image is larger than default)
            # Convert points to pixels for comparison (1 point ≈ 1.33 pixels)
            max_width_pixels = int(default_width * 1.33)
            max_height_pixels = int(default_height * 1.33)

            if (
                original_width <= max_width_pixels
                and original_height <= max_height_pixels
            ):
                # Image is small enough, use original size (copy to new image to avoid fp issues)
                self._log_info(
                    f"Using original image size: {original_width}x{original_height}px"
                )
                # Create a new image to avoid file pointer issues
                new_img = pil_img.copy()
                return new_img, (original_width, original_height)
            else:
                # Resize to default size while maintaining aspect ratio
                new_img = pil_img.copy()
                new_img.thumbnail(
                    (max_width_pixels, max_height_pixels), PILImage.Resampling.LANCZOS
                )
                new_size = new_img.size
                self._log_info(
                    f"Resized image from {original_width}x{original_height}px to {new_size[0]}x{new_size[1]}px"
                )
                return new_img, new_size

        except Exception as e:
            self._log_error(f"Image processing failed: {e}")
            # Return copy of original image if processing fails
            return pil_img.copy(), pil_img.size

    def _get_image_dimensions_lightweight(self, image_path_or_bytes):
        """Get image dimensions without full PIL processing to avoid corruption."""
        try:
            if isinstance(image_path_or_bytes, str):
                # File path
                with open(image_path_or_bytes, "rb") as f:
                    image_bytes = f.read()
            else:
                # Already bytes
                image_bytes = image_path_or_bytes

            # Use PIL but only for dimensions, then immediately close
            temp_img = PILImage.open(io.BytesIO(image_bytes))
            dimensions = temp_img.size
            temp_img.close()  # Immediately close to avoid corruption

            return dimensions

        except Exception as e:
            self._log_warning(f"Failed to get image dimensions: {e}")
            # Return default dimensions if detection fails
            return (100, 75)  # Default fallback

    def _adjust_row_height(self, sheet, row, image_height_pixels):
        """Adjust row height to accommodate image."""
        try:
            # Convert pixels to Excel points (1 point ≈ 1.33 pixels)
            required_height_points = max(15, image_height_pixels / 1.33)

            # Get current row height (default is ~15 points)
            current_height = sheet.row_dimensions[row].height or 15

            # Only increase height if necessary
            if required_height_points > current_height:
                sheet.row_dimensions[row].height = required_height_points
                self._log_info(
                    f"Adjusted row {row} height from {current_height:.1f} to {required_height_points:.1f} points"
                )
            else:
                self._log_info(
                    f"Row {row} height sufficient: {current_height:.1f} points (needed: {required_height_points:.1f})"
                )

        except Exception as e:
            self._log_warning(f"Failed to adjust row height for row {row}: {e}")

    def _adjust_column_width(self, sheet, col, image_width_pixels):
        """Adjust column width to accommodate image."""
        try:
            # Convert pixels to Excel character widths (rough approximation)
            # Excel default column width is ~8.43 characters ≈ 64 points ≈ 85 pixels
            required_width_chars = max(
                8.43, image_width_pixels / 10
            )  # ~10 pixels per character

            # Get current column width (default is ~8.43 characters)
            col_letter = get_column_letter(col)
            current_width = sheet.column_dimensions[col_letter].width or 8.43

            # Only increase width if necessary
            if required_width_chars > current_width:
                sheet.column_dimensions[col_letter].width = required_width_chars
                self._log_info(
                    f"Adjusted column {col_letter} width from {current_width:.1f} to {required_width_chars:.1f} characters"
                )
            else:
                self._log_info(
                    f"Column {col_letter} width sufficient: {current_width:.1f} characters (needed: {required_width_chars:.1f})"
                )

        except Exception as e:
            self._log_warning(f"Failed to adjust column width for column {col}: {e}")

    def _is_cell_address(self, mapping_key: str) -> bool:
        """Check if mapping key is a cell address (e.g., 'B14')."""
        return bool(re.match(r"^[A-Z]+\d+$", mapping_key.upper()))

    def _is_column_letter(self, mapping_key: str) -> bool:
        """Check if mapping key is a column letter (e.g., 'A', 'BC')."""
        # Excel columns are max 3 letters (up to XFD = 16384 columns)
        # Common valid patterns: A, B, Z, AA, AB, ZZ, AAA, XFD
        if not mapping_key or len(mapping_key) > 3:
            return False

        # Must be all uppercase letters
        if not re.match(r"^[A-Z]+$", mapping_key.upper()):
            return False

        # Try to convert to column index - if it fails, not a valid column
        try:
            from openpyxl.utils import column_index_from_string

            column_index_from_string(mapping_key.upper())
            return True
        except ValueError:
            return False

    def _find_header_column(
        self,
        sheet: Worksheet,
        header_row: int,
        header_text: str,
        config: Dict[str, Any],
        base_col: int = 1,
    ) -> Optional[int]:
        """Find column index for header text."""
        max_columns = config.get("max_columns", 20)
        start_col = base_col

        for col_offset in range(max_columns):
            col_idx = start_col + col_offset
            cell = sheet.cell(row=header_row, column=col_idx)

            if cell.value and header_text.lower() in str(cell.value).lower():
                return col_idx

        return None

    def _find_key_row(
        self,
        sheet: Worksheet,
        base_row: int,
        base_col: int,
        key_text: str,
        config: Dict[str, Any],
    ) -> Optional[int]:
        """Find row index for key text in vertical layout."""
        max_rows = config.get("max_rows", 20)
        start_row = base_row

        for row_offset in range(max_rows):
            row_idx = start_row + row_offset
            cell = sheet.cell(row=row_idx, column=base_col)

            if cell.value and key_text.lower() in str(cell.value).lower():
                return row_idx

        return None

    def _parse_cell_address(self, cell_address: str) -> Tuple[int, int]:
        """Parse cell address into row and column."""
        try:
            col_letter, row = coordinate_from_string(cell_address)
            col_idx = column_index_from_string(col_letter)
            return row, col_idx
        except Exception:
            raise ValueError(f"Invalid cell address: {cell_address}")

    def _detect_all_subtables(
        self, sheet: Worksheet, sheet_config: Dict[str, Any]
    ) -> Dict[str, Dict[str, Any]]:
        """
        Detect all subtables before making any modifications.
        
        Returns:
            Dictionary mapping subtable names to their detection info:
            {
                "subtable_name": {
                    "location": {"row": int, "col": int, "found": bool},
                    "type": str,
                    "priority": int,
                    "config": dict
                }
            }
        """
        detected_subtables = {}
        
        for subtable_config in sheet_config.get("subtables", []):
            subtable_name = subtable_config["name"]
            subtable_type = subtable_config["type"]
            
            # Skip if no data_update configuration
            if "data_update" not in subtable_config:
                self._log_info(
                    f"Skipping detection for '{subtable_name}' - no data_update configuration"
                )
                continue
            
            # Find starting location
            location = self._find_update_location(sheet, subtable_config["header_search"])
            
            if location["found"]:
                # Determine processing priority based on type
                priority = self._determine_processing_order(subtable_type)
                
                detected_subtables[subtable_name] = {
                    "location": location,
                    "type": subtable_type,
                    "priority": priority,
                    "config": subtable_config
                }
                self._log_info(
                    f"Detected subtable '{subtable_name}' at {location['address']} "
                    f"(type: {subtable_type}, priority: {priority})"
                )
            else:
                self._log_error(f"Could not detect location for subtable '{subtable_name}'")
                # Still add to detected list with not found status
                detected_subtables[subtable_name] = {
                    "location": location,
                    "type": subtable_type,
                    "priority": 999,  # Low priority for not found
                    "config": subtable_config
                }
        
        return detected_subtables

    def _determine_processing_order(self, subtable_type: str) -> int:
        """
        Determine processing priority based on subtable type.
        
        Priority 1: Fixed-size tables (process first)
        Priority 2: Expandable tables (process after fixed tables)
        
        Args:
            subtable_type: Type of subtable (key_value_pairs, matrix_table, table)
            
        Returns:
            Priority number (lower = higher priority)
        """
        if subtable_type in ["key_value_pairs", "matrix_table"]:
            return 1  # Fixed-size tables
        elif subtable_type == "table":
            return 2  # Expandable tables
        else:
            return 3  # Unknown types (process last)

    def _process_subtables_in_order(
        self,
        sheet: Worksheet,
        detected_subtables: Dict[str, Dict[str, Any]],
        update_data: Dict[str, Any],
    ) -> None:
        """
        Process subtables in correct order with position tracking.
        
        Args:
            sheet: Excel worksheet
            detected_subtables: Dictionary of detected subtables with metadata
            update_data: Data to update in the subtables
        """
        # Sort subtables by priority (lower number = higher priority)
        sorted_subtables = sorted(
            detected_subtables.items(),
            key=lambda x: (x[1]["priority"], x[0])  # Sort by priority, then name
        )
        
        # Track cumulative row shifts for position adjustment
        cumulative_row_shift = 0
        
        for subtable_name, subtable_info in sorted_subtables:
            if not subtable_info["location"]["found"]:
                self._log_warning(f"Skipping '{subtable_name}' - location not found")
                continue
            
            # Skip if no data provided for this subtable
            if subtable_name not in update_data:
                self._log_warning(f"No data provided for subtable '{subtable_name}'")
                continue
            
            # Adjust location based on cumulative shifts from previous updates
            adjusted_location = self._adjust_location_for_shifts(
                subtable_info["location"], cumulative_row_shift
            )
            
            self._log_info(
                f"Processing subtable '{subtable_name}' at adjusted location "
                f"row={adjusted_location['row']} (shift={cumulative_row_shift})"
            )
            
            # Update the subtable and get expansion amount
            expansion_rows = self._update_subtable_with_expansion(
                sheet,
                subtable_info["config"],
                update_data[subtable_name],
                adjusted_location,
                cumulative_row_shift
            )
            
            # Update cumulative shift for next subtables
            if expansion_rows != 0:
                cumulative_row_shift += expansion_rows
                self._log_info(
                    f"Subtable '{subtable_name}' expanded by {expansion_rows} rows. "
                    f"Total shift: {cumulative_row_shift}"
                )

    def _adjust_location_for_shifts(
        self, original_location: Dict[str, Any], cumulative_shift: int
    ) -> Dict[str, Any]:
        """
        Adjust subtable location based on previous expansions.
        
        Args:
            original_location: Original location dictionary
            cumulative_shift: Number of rows to shift down
            
        Returns:
            Adjusted location dictionary
        """
        adjusted_location = original_location.copy()
        
        if cumulative_shift != 0:
            # Adjust row position
            adjusted_location["row"] = original_location["row"] + cumulative_shift
            
            # Update cell address if present
            if "address" in original_location and original_location["address"]:
                col_letter = get_column_letter(original_location["col"])
                adjusted_location["address"] = f"{col_letter}{adjusted_location['row']}"
        
        return adjusted_location

    def _detect_table_boundaries(
        self,
        sheet: Worksheet,
        data_start_row: int,
        data_start_col: int,
        column_count: int,
        max_rows: int = 1000,
    ) -> Tuple[int, bool, Optional[range]]:
        """
        Detect where table ends using consecutive empty row logic.
        
        Args:
            sheet: Excel worksheet
            data_start_row: Row where table data starts
            data_start_col: Column where table starts
            column_count: Number of columns in the table
            max_rows: Maximum rows to scan
            
        Returns:
            Tuple of (table_end_row, content_exists_below, affected_columns_range)
        """
        consecutive_empty_rows = 0
        table_end_row = data_start_row
        
        # Scan rows to find table end
        for row_offset in range(max_rows):
            current_row = data_start_row + row_offset
            
            # Check if we've gone beyond the sheet bounds
            if current_row > sheet.max_row:
                break
            
            # Check if entire row is empty
            row_is_empty = True
            for col_offset in range(column_count):
                cell = sheet.cell(row=current_row, column=data_start_col + col_offset)
                if cell.value is not None and str(cell.value).strip():
                    row_is_empty = False
                    break
            
            if row_is_empty:
                consecutive_empty_rows += 1
                # Stop after 2-3 consecutive empty rows (matching extract logic)
                if consecutive_empty_rows >= 2:
                    table_end_row = current_row - consecutive_empty_rows
                    break
            else:
                consecutive_empty_rows = 0
                table_end_row = current_row
        
        # Check if there's content below the table
        content_exists_below = False
        # If we found empty rows, skip them; otherwise check immediately after table end
        first_content_row = table_end_row + max(consecutive_empty_rows, 1) + 1
        
        # Scan a reasonable range below table to check for content
        for row in range(first_content_row, min(first_content_row + 50, sheet.max_row + 1)):
            for col in sheet[row]:
                if col.value is not None and str(col.value).strip():
                    content_exists_below = True
                    break
            if content_exists_below:
                break
        
        # Define affected columns range
        affected_columns = range(data_start_col, data_start_col + column_count)
        
        self._log_info(
            f"Table boundaries: ends at row {table_end_row}, "
            f"content below: {content_exists_below}, "
            f"columns {data_start_col}-{data_start_col + column_count - 1}"
        )
        
        return table_end_row, content_exists_below, affected_columns

    def _preserve_content_below_table(
        self,
        sheet: Worksheet,
        table_end_row: int,
        affected_columns: range,
        preserve_to_row: Optional[int] = None,
    ) -> Dict[str, Any]:
        """
        Extract all content below table for later restoration.
        Preserves entire rows to maintain document structure.
        
        Args:
            sheet: Excel worksheet
            table_end_row: Last row of the table
            affected_columns: Range of columns affected by the table (kept for API compatibility)
            preserve_to_row: Optional end row for preservation (default: sheet.max_row)
            
        Returns:
            Dictionary containing preserved content from all columns
        """
        preserved_content = {
            "cells": {},
            "merged_cells": [],
            "row_heights": {},
            "images": []
        }
        
        start_row = table_end_row + 1
        end_row = preserve_to_row or sheet.max_row
        
        # Preserve cell data and formatting for entire rows
        for row in range(start_row, end_row + 1):
            # Preserve all columns in the row, not just affected columns
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                
                # Skip empty cells
                if cell.value is None and not cell.has_style:
                    continue
                
                cell_data = {
                    "value": cell.value,
                    "data_type": cell.data_type,
                }
                
                # Preserve formula if present
                if cell.data_type == 'f':
                    cell_data["formula"] = cell.value
                
                # Preserve formatting
                if cell.has_style:
                    cell_data["font"] = copy.copy(cell.font)
                    cell_data["fill"] = copy.copy(cell.fill)
                    cell_data["border"] = copy.copy(cell.border)
                    cell_data["alignment"] = copy.copy(cell.alignment)
                    cell_data["number_format"] = cell.number_format
                
                cell_address = f"{get_column_letter(col)}{row}"
                preserved_content["cells"][cell_address] = cell_data
        
        # Preserve merged cells in affected rows (all columns)
        for merged_range in sheet.merged_cells.ranges:
            if merged_range.min_row >= start_row:
                preserved_content["merged_cells"].append({
                    "range": str(merged_range),
                    "min_row": merged_range.min_row,
                    "max_row": merged_range.max_row,
                    "min_col": merged_range.min_col,
                    "max_col": merged_range.max_col
                })
        
        # Preserve row heights
        for row in range(start_row, end_row + 1):
            if sheet.row_dimensions[row].height:
                preserved_content["row_heights"][row] = sheet.row_dimensions[row].height
        
        # Preserve images in affected area
        if hasattr(sheet, '_images'):
            for img in sheet._images:
                # Check if image is in preserved area
                if hasattr(img, 'anchor') and img.anchor:
                    # Extract position from anchor
                    cell_ref = str(img.anchor).split('!')[0] if '!' in str(img.anchor) else str(img.anchor)
                    if cell_ref:
                        try:
                            img_col, img_row = coordinate_from_string(cell_ref)
                            img_col_idx = column_index_from_string(img_col)
                            if img_row >= start_row:
                                preserved_content["images"].append({
                                    "image": img,
                                    "anchor": str(img.anchor),
                                    "row": img_row,
                                    "col": img_col_idx
                                })
                        except:
                            pass  # Skip if we can't parse the anchor
        
        self._log_info(
            f"Preserved content: {len(preserved_content['cells'])} cells, "
            f"{len(preserved_content['merged_cells'])} merged ranges, "
            f"{len(preserved_content['images'])} images"
        )
        
        return preserved_content

    def _restore_preserved_content(
        self,
        sheet: Worksheet,
        preserved_content: Dict[str, Any],
        row_shift: int,
    ) -> None:
        """
        Restore preserved content at shifted positions.
        
        Args:
            sheet: Excel worksheet
            preserved_content: Dictionary with preserved content
            row_shift: Number of rows to shift content down
        """
        if row_shift == 0:
            self._log_info("No row shift needed, skipping content restoration")
            return
        
        self._log_info(f"Restoring preserved content with row shift of {row_shift}")
        
        # First, unmerge any cells that will be affected
        for merged_info in preserved_content["merged_cells"]:
            try:
                sheet.unmerge_cells(merged_info["range"])
            except:
                pass  # Cell might already be unmerged
        
        # Restore cell data with shifted positions
        for original_address, cell_data in preserved_content["cells"].items():
            # Parse original position
            col_letter, row = coordinate_from_string(original_address)
            col_idx = column_index_from_string(col_letter)
            
            # Calculate new position
            new_row = row + row_shift
            new_address = f"{col_letter}{new_row}"
            
            # Get target cell
            target_cell = sheet.cell(row=new_row, column=col_idx)
            
            # Restore value (with formula reference updates if needed)
            if cell_data.get("data_type") == 'f' and cell_data.get("formula"):
                # Update formula references
                updated_formula = self._update_formula_references(
                    cell_data["formula"], row_shift
                )
                target_cell.value = updated_formula
                target_cell.data_type = 'f'
            else:
                target_cell.value = cell_data["value"]
                if "data_type" in cell_data:
                    target_cell.data_type = cell_data["data_type"]
            
            # Restore formatting
            if "font" in cell_data:
                target_cell.font = cell_data["font"]
            if "fill" in cell_data:
                target_cell.fill = cell_data["fill"]
            if "border" in cell_data:
                target_cell.border = cell_data["border"]
            if "alignment" in cell_data:
                target_cell.alignment = cell_data["alignment"]
            if "number_format" in cell_data:
                target_cell.number_format = cell_data["number_format"]
        
        # Re-merge cells at new positions
        for merged_info in preserved_content["merged_cells"]:
            new_min_row = merged_info["min_row"] + row_shift
            new_max_row = merged_info["max_row"] + row_shift
            
            new_range = f"{get_column_letter(merged_info['min_col'])}{new_min_row}:" \
                       f"{get_column_letter(merged_info['max_col'])}{new_max_row}"
            
            try:
                sheet.merge_cells(new_range)
            except Exception as e:
                self._log_warning(f"Failed to merge cells {new_range}: {e}")
        
        # Restore row heights
        for original_row, height in preserved_content["row_heights"].items():
            new_row = original_row + row_shift
            sheet.row_dimensions[new_row].height = height
        
        # Restore images at new positions
        for img_info in preserved_content["images"]:
            try:
                # Update anchor position
                new_row = img_info["row"] + row_shift
                new_anchor = f"{get_column_letter(img_info['col'])}{new_row}"
                
                # Re-add image with new anchor
                img = img_info["image"]
                img.anchor = new_anchor
                # Note: Image might already be in sheet._images, just updating anchor
                
            except Exception as e:
                self._log_warning(f"Failed to reposition image: {e}")
        
        self._log_info(
            f"Content restoration complete: {len(preserved_content['cells'])} cells, "
            f"{len(preserved_content['merged_cells'])} merged ranges"
        )

    def _update_formula_references(self, formula: str, row_shift: int) -> str:
        """
        Update cell references in formula to account for row shift.
        
        Args:
            formula: Original formula string
            row_shift: Number of rows to shift references
            
        Returns:
            Updated formula with shifted references
        """
        # Pattern to match cell references (e.g., A1, $B$2, Sheet1!C3)
        cell_ref_pattern = r'(\$?[A-Z]+\$?)(\d+)'
        
        def update_reference(match):
            col_part = match.group(1)
            row_part = int(match.group(2))
            
            # Don't update absolute row references (with $)
            if '$' in match.group(0).split(col_part)[1]:
                return match.group(0)
            
            # Update row number
            new_row = row_part + row_shift
            return f"{col_part}{new_row}"
        
        # Update all cell references in the formula
        updated_formula = re.sub(cell_ref_pattern, update_reference, formula)
        
        return updated_formula

    def _update_subtable_with_expansion(
        self,
        sheet: Worksheet,
        subtable_config: Dict[str, Any],
        data: Any,
        adjusted_location: Dict[str, Any],
        cumulative_shift: int,
    ) -> int:
        """
        Update subtable with support for unlimited row expansion.
        
        Args:
            sheet: Excel worksheet
            subtable_config: Subtable configuration
            data: Data to update
            adjusted_location: Location adjusted for previous shifts
            cumulative_shift: Total shift from previous expansions
            
        Returns:
            Number of rows expanded (positive) or contracted (negative)
        """
        subtable_name = subtable_config["name"]
        subtable_type = subtable_config["type"]
        update_config = subtable_config["data_update"]
        
        # Apply offsets to starting location
        base_row = adjusted_location["row"] + update_config.get("headers_row_offset", 0)
        base_col = adjusted_location["col"] + update_config.get("headers_col_offset", 0)
        
        # For table type, check if expansion is needed
        if subtable_type == "table" and isinstance(data, list):
            # Get table configuration
            data_row_offset = update_config.get("data_row_offset", 1)
            data_start_row = base_row + data_row_offset
            
            # Detect current table boundaries
            column_count = len(update_config.get("column_mappings", {}))
            max_rows = update_config.get("max_rows", 1000)
            
            self._log_info(
                f"Detecting boundaries for table at row {data_start_row}, col {base_col}, "
                f"with {column_count} columns"
            )
            
            table_end_row, content_below, affected_columns = self._detect_table_boundaries(
                sheet, data_start_row, base_col, column_count, max_rows
            )
            
            # Calculate expansion needed
            current_table_rows = table_end_row - data_start_row + 1
            new_data_rows = len(data)
            expansion_rows = new_data_rows - current_table_rows
            
            self._log_info(
                f"Table '{subtable_name}': current rows={current_table_rows}, "
                f"new rows={new_data_rows}, expansion={expansion_rows}"
            )
            
            # Handle expansion if needed
            if expansion_rows > 0 and content_below:
                # Check expansion limits
                max_expansion = update_config.get("max_expansion_rows", 1000)
                if expansion_rows > max_expansion:
                    self._log_error(
                        f"Expansion of {expansion_rows} rows exceeds limit of {max_expansion}"
                    )
                    expansion_rows = max_expansion
                
                # Preserve content below table
                preserved_content = self._preserve_content_below_table(
                    sheet, table_end_row, affected_columns
                )
                
                # Clear entire rows in the expansion area (where preserved content currently is)
                clear_start_row = table_end_row + 1
                clear_end_row = table_end_row + expansion_rows
                
                # First, unmerge any cells in the area we're about to clear
                merged_ranges_to_remove = []
                for merged_range in sheet.merged_cells.ranges:
                    if (merged_range.min_row >= clear_start_row and 
                        merged_range.min_row <= clear_end_row):
                        merged_ranges_to_remove.append(str(merged_range))
                
                for range_str in merged_ranges_to_remove:
                    try:
                        sheet.unmerge_cells(range_str)
                    except:
                        pass  # Ignore if already unmerged
                
                # Now clear all columns in the rows
                for row in range(clear_start_row, clear_end_row + 1):
                    for col in range(1, sheet.max_column + 1):
                        try:
                            sheet.cell(row=row, column=col).value = None
                        except AttributeError:
                            # Skip if it's a merged cell that we couldn't unmerge
                            pass
                
                # Update the table with new data
                self._update_table_with_offsets(
                    sheet, base_row, base_col, update_config, data
                )
                
                # Restore preserved content at new positions
                self._restore_preserved_content(sheet, preserved_content, expansion_rows)
                
                return expansion_rows
            else:
                # No expansion needed or no content below - update normally
                self._update_table_with_offsets(
                    sheet, base_row, base_col, update_config, data
                )
                return 0
        else:
            # Non-expandable table types - update normally
            if subtable_type == "key_value_pairs":
                self._update_key_value_pairs_with_offsets(
                    sheet, base_row, base_col, update_config, data
                )
            elif subtable_type == "matrix_table":
                self._update_matrix_table_with_offsets(
                    sheet, base_row, base_col, update_config, data
                )
            else:
                # Default table update (no expansion)
                self._update_table_with_offsets(
                    sheet, base_row, base_col, update_config, data
                )
            
            return 0  # No expansion for fixed-size tables

    def _validate_update_config(self, config: Dict[str, Any]) -> None:
        """Validate update configuration structure."""
        if "sheet_configs" not in config:
            raise ValidationError("Configuration must contain 'sheet_configs'")

        for sheet_name, sheet_config in config.get("sheet_configs", {}).items():
            if "subtables" not in sheet_config:
                raise ValidationError(f"Sheet '{sheet_name}' must contain 'subtables'")

            for subtable in sheet_config.get("subtables", []):
                # Validate required fields (data_update OR data_extraction)
                required_base_fields = ["name", "type", "header_search"]
                for field in required_base_fields:
                    if field not in subtable:
                        raise ValidationError(
                            f"Subtable missing required field: {field}"
                        )

                # Must have either data_update or data_extraction
                if "data_update" not in subtable and "data_extraction" not in subtable:
                    raise ValidationError(
                        f"Subtable '{subtable['name']}' must have either 'data_update' or 'data_extraction'"
                    )

                # If using data_extraction, warn and skip
                if "data_extraction" in subtable and "data_update" not in subtable:
                    logger.warning(
                        f"Subtable '{subtable['name']}' has data_extraction but not data_update - skipping"
                    )

                # Validate header search method
                header_search = subtable.get("header_search", {})
                method = header_search.get("method")

                if method == "cell_address":
                    if "cell" not in header_search:
                        raise ValidationError(
                            f"cell_address method requires 'cell' field in {subtable['name']}"
                        )
                elif method == "contains_text":
                    required = ["text", "column", "search_range"]
                    for field in required:
                        if field not in header_search:
                            raise ValidationError(
                                f"contains_text method requires '{field}' field in {subtable['name']}"
                            )
                else:
                    raise ValidationError(
                        f"Unsupported method '{method}' in {subtable['name']}"
                    )

                # Validate data_update structure if present
                if "data_update" in subtable:
                    data_update = subtable.get("data_update", {})
                    if "column_mappings" not in data_update:
                        raise ValidationError(
                            f"column_mappings required in data_update for {subtable['name']}"
                        )
                    
                    # Validate new optional fields
                    if "max_expansion_rows" in data_update:
                        max_exp = data_update["max_expansion_rows"]
                        if not isinstance(max_exp, int) or max_exp <= 0:
                            raise ValidationError(
                                f"max_expansion_rows must be a positive integer in {subtable['name']}"
                            )
                    
                    if "expansion_behavior" in data_update:
                        valid_behaviors = ["preserve_below", "overwrite", "error"]
                        if data_update["expansion_behavior"] not in valid_behaviors:
                            raise ValidationError(
                                f"expansion_behavior must be one of {valid_behaviors} in {subtable['name']}"
                            )
                    
                    if "copy_first_row_style" in data_update:
                        if not isinstance(data_update["copy_first_row_style"], bool):
                            raise ValidationError(
                                f"copy_first_row_style must be a boolean in {subtable['name']}"
                            )

    def _add_update_log_sheet(self) -> None:
        """Create diagnostic sheet with update results."""
        try:
            # Remove existing log sheet if present
            if "update_log" in self.workbook.sheetnames:
                del self.workbook["update_log"]

            # Create new log sheet
            log_sheet = self.workbook.create_sheet("update_log")

            # Add headers
            headers = [
                "Timestamp",
                "Operation",
                "Cell/Range",
                "Status",
                "Details",
                "Original Value",
                "New Value",
            ]
            for col_idx, header in enumerate(headers, 1):
                log_sheet.cell(row=1, column=col_idx, value=header)

            # Add log entries
            for row_idx, log_entry in enumerate(self.update_log, 2):
                log_sheet.cell(row=row_idx, column=1, value=log_entry["timestamp"])
                log_sheet.cell(row=row_idx, column=2, value=log_entry["operation"])
                log_sheet.cell(row=row_idx, column=3, value=log_entry.get("cell", ""))
                log_sheet.cell(row=row_idx, column=4, value=log_entry["status"])
                log_sheet.cell(row=row_idx, column=5, value=log_entry["details"])
                log_sheet.cell(
                    row=row_idx, column=6, value=log_entry.get("original_value", "")
                )
                log_sheet.cell(
                    row=row_idx, column=7, value=log_entry.get("new_value", "")
                )

            self._log_info(
                f"Created update_log sheet with {len(self.update_log)} entries"
            )

        except Exception as e:
            logger.warning(f"Failed to create update log sheet: {e}")

    def _save_updated_workbook(self) -> str:
        """Save updated workbook to new file."""
        try:
            # Create output path
            base_name = os.path.splitext(os.path.basename(self.file_path))[0]
            output_dir = os.path.dirname(self.file_path)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(
                output_dir, f"updated_{base_name}_{timestamp}.xlsx"
            )

            # Save workbook with improved error handling
            try:
                self.workbook.save(output_path)
                return output_path
            except Exception as e:
                # Check for I/O operation on closed file error
                if "I/O operation on closed file" in str(e):
                    self._log_error(f"I/O operation on closed file error: {e}")
                    self._log_error("This suggests the workbook has corrupted file handles")
                    # Try to reload and re-save the workbook
                    try:
                        self._log_info("Attempting workbook reload and re-save...")
                        fresh_workbook = load_workbook(self.file_path)
                        fresh_workbook.save(output_path)
                        self._log_info("Successfully saved using fresh workbook")
                        return output_path
                    except Exception as reload_error:
                        self._log_error(f"Workbook reload also failed: {reload_error}")
                        raise e
                else:
                    raise e
        finally:
            # Clean up temporary image files after workbook save
            self._cleanup_temp_files()

    def _log_info(self, message: str) -> None:
        """Log info message."""
        entry = {
            "timestamp": datetime.now().isoformat(),
            "operation": "INFO",
            "status": "INFO",
            "details": message,
        }
        self.update_log.append(entry)
        logger.info(message)

    def _log_success(
        self,
        message: str,
        cell: str = "",
        original_value: Any = "",
        new_value: Any = "",
    ) -> None:
        """Log success message."""
        entry = {
            "timestamp": datetime.now().isoformat(),
            "operation": "UPDATE",
            "cell": cell,
            "status": "SUCCESS",
            "details": message,
            "original_value": str(original_value) if original_value is not None else "",
            "new_value": str(new_value) if new_value is not None else "",
        }
        self.update_log.append(entry)
        logger.info(message)

    def _log_warning(self, message: str, cell: str = "") -> None:
        """Log warning message."""
        entry = {
            "timestamp": datetime.now().isoformat(),
            "operation": "WARNING",
            "cell": cell,
            "status": "WARNING",
            "details": message,
        }
        self.update_log.append(entry)
        logger.warning(message)

    def _log_error(self, message: str, cell: str = "") -> None:
        """Log error message."""
        entry = {
            "timestamp": datetime.now().isoformat(),
            "operation": "ERROR",
            "cell": cell,
            "status": "ERROR",
            "details": message,
        }
        self.update_log.append(entry)
        logger.error(message)

    def _log_existing_content(self) -> None:
        """Log all existing content for preservation verification."""
        try:
            total_images = 0
            total_charts = 0
            total_data_cells = 0

            for sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]

                # Count images
                sheet_images = 0
                if hasattr(sheet, "_images") and sheet._images:
                    sheet_images = len(sheet._images)
                    total_images += sheet_images

                # Count charts
                sheet_charts = 0
                if hasattr(sheet, "_charts") and sheet._charts:
                    sheet_charts = len(sheet._charts)
                    total_charts += sheet_charts

                # Count data cells (non-empty cells)
                sheet_data_cells = 0
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            sheet_data_cells += 1
                total_data_cells += sheet_data_cells

                self._log_info(
                    f"Sheet '{sheet_name}': {sheet_images} images, {sheet_charts} charts, {sheet_data_cells} data cells"
                )

            self._log_info(
                f"Total content: {total_images} images, {total_charts} charts, {total_data_cells} data cells"
            )

        except Exception as e:
            self._log_warning(f"Failed to log existing content: {e}")

    def _verify_sheet_preservation(self, sheet_name: str) -> None:
        """Verify that non-updated sheet content is preserved."""
        try:
            sheet = self.workbook[sheet_name]

            # Count current content
            images = (
                len(sheet._images) if hasattr(sheet, "_images") and sheet._images else 0
            )
            charts = (
                len(sheet._charts) if hasattr(sheet, "_charts") and sheet._charts else 0
            )

            # Count non-empty cells
            data_cells = 0
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None and str(cell.value) != "!#VALUE":
                        data_cells += 1

            self._log_info(
                f"Verified preservation of sheet '{sheet_name}': {images} images, {charts} charts, {data_cells} data cells"
            )

            # Log any #VALUE errors found (indicating potential data loss)
            value_errors = 0
            for row in sheet.iter_rows():
                for cell in row:
                    if str(cell.value) == "!#VALUE":
                        value_errors += 1
                        self._log_warning(
                            f"Found !#VALUE error in {sheet_name}:{cell.coordinate}"
                        )

            if value_errors > 0:
                self._log_error(
                    f"Sheet '{sheet_name}' has {value_errors} !#VALUE errors - possible data loss"
                )

        except Exception as e:
            self._log_warning(
                f"Failed to verify sheet preservation for '{sheet_name}': {e}"
            )

    def _cleanup_temp_files(self) -> None:
        """Clean up temporary image files."""
        for temp_file in self.temp_image_files:
            try:
                if os.path.exists(temp_file):
                    os.unlink(temp_file)
            except Exception as e:
                logger.warning(f"Failed to cleanup temp file {temp_file}: {e}")
        self.temp_image_files.clear()

    def close(self) -> None:
        """Close workbook and cleanup resources."""
        self._cleanup_temp_files()
        if hasattr(self, "workbook"):
            self.workbook.close()

    def delete_sheets(self, sheet_names: List[str], include_update_log: bool = False) -> str:
        """
        Delete specified sheets from the workbook.

        Args:
            sheet_names: List of sheet names to delete
            include_update_log: Whether to include diagnostic update_log sheet

        Returns:
            Path to updated Excel file
        """
        try:
            self._log_info(f"Starting sheet deletion process for sheets: {sheet_names}")

            # Validate that at least one sheet will remain
            current_sheets = list(self.workbook.sheetnames)
            sheets_to_delete = [name for name in sheet_names if name in current_sheets]
            sheets_not_found = [name for name in sheet_names if name not in current_sheets]
            
            if sheets_not_found:
                self._log_warning(f"Sheets not found in workbook: {sheets_not_found}")
            
            if not sheets_to_delete:
                self._log_error("No valid sheets to delete")
                raise ValidationError("None of the specified sheets exist in the workbook")
            
            remaining_sheets = len(current_sheets) - len(sheets_to_delete)
            if remaining_sheets < 1:
                self._log_error("Cannot delete all sheets - at least one sheet must remain")
                raise ValidationError("Cannot delete all sheets from workbook. At least one sheet must remain.")
            
            # Delete the sheets
            for sheet_name in sheets_to_delete:
                self._log_info(f"Deleting sheet: {sheet_name}")
                del self.workbook[sheet_name]
                self.update_log.append({
                    "timestamp": datetime.now().isoformat(),
                    "operation": "delete_sheet",
                    "cell": sheet_name,
                    "status": "success",
                    "details": f"Deleted sheet: {sheet_name}"
                })
            
            self._log_info(f"Successfully deleted {len(sheets_to_delete)} sheets")
            
            # Add diagnostic log sheet if requested
            if include_update_log:
                self._add_update_log_sheet()
            
            # Save updated workbook
            output_path = self._save_updated_workbook()
            self._log_info(f"Sheet deletion completed successfully: {output_path}")
            
            return output_path
            
        except Exception as e:
            self._log_error(f"Sheet deletion failed: {e}")
            raise ExcelProcessingError(f"Failed to delete sheets: {e}")

    def add_sheets(self, source_file_path: str, sheet_names: List[str], include_update_log: bool = False, sheet_positions: Optional[Dict[str, Optional[int]]] = None, sheet_replace_flags: Optional[Dict[str, Optional[bool]]] = None) -> str:
        """
        Add sheets from source workbook to this workbook.

        Args:
            source_file_path: Path to source Excel file
            sheet_names: List of sheet names to copy from source
            include_update_log: Whether to include diagnostic update_log sheet
            sheet_positions: Optional dict mapping sheet names to their desired positions
            sheet_replace_flags: Optional dict mapping sheet names to replace behavior (True/False/None)

        Returns:
            Path to updated Excel file
        """
        try:
            self._log_info(f"Starting sheet addition process from {source_file_path}")
            self._log_info(f"Sheets to add: {sheet_names}")
            if sheet_positions:
                self._log_info(f"Sheet positions: {sheet_positions}")
            if sheet_replace_flags:
                self._log_info(f"Sheet replace flags: {sheet_replace_flags}")
            
            # Load source workbook
            source_workbook = load_workbook(source_file_path, data_only=False)
            
            try:
                # Validate source sheets exist
                source_sheets = source_workbook.sheetnames
                sheets_to_copy = [name for name in sheet_names if name in source_sheets]
                sheets_not_found = [name for name in sheet_names if name not in source_sheets]
                
                if sheets_not_found:
                    self._log_warning(f"Sheets not found in source workbook: {sheets_not_found}")
                
                if not sheets_to_copy:
                    self._log_error("No valid sheets to copy")
                    raise ValidationError("None of the specified sheets exist in the source workbook")
                
                # Check for duplicate sheet names
                current_sheets = self.workbook.sheetnames
                duplicate_sheets = [name for name in sheets_to_copy if name in current_sheets]
                
                # Initialize collections for different handling strategies
                rename_map = {}
                sheets_to_skip = []
                sheets_to_replace = []
                sheets_to_rename = []
                
                # Process duplicates based on replace flags
                for sheet_name in duplicate_sheets:
                    replace_flag = sheet_replace_flags.get(sheet_name) if sheet_replace_flags else None
                    
                    if replace_flag is True:
                        sheets_to_replace.append(sheet_name)
                    elif replace_flag is False:
                        sheets_to_skip.append(sheet_name)
                    else:
                        # Default behavior - rename
                        sheets_to_rename.append(sheet_name)
                
                # Handle replacements first
                for sheet_name in sheets_to_replace:
                    # Store original position if no new position specified
                    original_position = list(self.workbook.sheetnames).index(sheet_name)
                    del self.workbook[sheet_name]
                    self._log_info(f"Deleted existing sheet '{sheet_name}' for replacement")
                    
                    # Update position if not specified
                    if sheet_positions is None:
                        sheet_positions = {}
                    if sheet_positions.get(sheet_name) is None:
                        sheet_positions[sheet_name] = original_position
                        self._log_info(f"Will insert replacement sheet '{sheet_name}' at original position {original_position}")
                
                # Skip sheets marked as no-replace
                for sheet_name in sheets_to_skip:
                    sheets_to_copy.remove(sheet_name)
                    self._log_warning(f"Skipping sheet '{sheet_name}' - already exists and replace=false")
                    self.update_log.append({
                        "timestamp": datetime.now().isoformat(),
                        "operation": "add_sheet_skipped",
                        "cell": sheet_name,
                        "status": "skipped",
                        "details": f"Sheet '{sheet_name}' already exists and replace=false"
                    })
                
                # Handle renames
                for sheet_name in sheets_to_rename:
                    counter = 1
                    new_name = f"{sheet_name}_{counter}"
                    while new_name in current_sheets or new_name in rename_map.values():
                        counter += 1
                        new_name = f"{sheet_name}_{counter}"
                    rename_map[sheet_name] = new_name
                    self._log_warning(f"Sheet '{sheet_name}' already exists, will be renamed to '{new_name}'")
                
                # Copy sheets
                for sheet_name in sheets_to_copy:
                    self._log_info(f"Copying sheet: {sheet_name}")
                    source_sheet = source_workbook[sheet_name]
                    
                    # Create new sheet in target workbook
                    target_sheet_name = rename_map.get(sheet_name, sheet_name)
                    
                    # Get position for this specific sheet
                    position = None
                    if sheet_positions:
                        position = sheet_positions.get(sheet_name)
                    
                    if position is not None:
                        target_sheet = self.workbook.create_sheet(target_sheet_name, position)
                        self._log_info(f"Created sheet '{target_sheet_name}' at position {position}")
                    else:
                        target_sheet = self.workbook.create_sheet(target_sheet_name)
                        self._log_info(f"Created sheet '{target_sheet_name}' at end")
                    
                    # Copy sheet properties
                    target_sheet.sheet_properties = copy.copy(source_sheet.sheet_properties)
                    
                    # Copy all cells with their properties
                    for row in source_sheet.iter_rows():
                        for cell in row:
                            target_cell = target_sheet[cell.coordinate]
                            
                            # Copy cell value
                            if cell.value is not None:
                                target_cell.value = cell.value
                            
                            # Copy cell style
                            if cell.has_style:
                                target_cell.font = copy.copy(cell.font)
                                target_cell.border = copy.copy(cell.border)
                                target_cell.fill = copy.copy(cell.fill)
                                target_cell.number_format = cell.number_format
                                target_cell.protection = copy.copy(cell.protection)
                                target_cell.alignment = copy.copy(cell.alignment)
                    
                    # Copy merged cells
                    for merged_range in source_sheet.merged_cells.ranges:
                        target_sheet.merge_cells(str(merged_range))
                    
                    # Copy column dimensions
                    for col_letter, col_dim in source_sheet.column_dimensions.items():
                        target_sheet.column_dimensions[col_letter].width = col_dim.width
                        target_sheet.column_dimensions[col_letter].hidden = col_dim.hidden
                    
                    # Copy row dimensions
                    for row_num, row_dim in source_sheet.row_dimensions.items():
                        target_sheet.row_dimensions[row_num].height = row_dim.height
                        target_sheet.row_dimensions[row_num].hidden = row_dim.hidden
                    
                    # Note: Images and charts would require additional handling
                    # For now, we'll log a warning if they exist
                    if hasattr(source_sheet, '_images') and source_sheet._images:
                        self._log_warning(f"Sheet '{sheet_name}' contains images which were not copied")
                    if hasattr(source_sheet, '_charts') and source_sheet._charts:
                        self._log_warning(f"Sheet '{sheet_name}' contains charts which were not copied")
                    
                    # Determine if this was a replacement
                    was_replaced = sheet_name in sheets_to_replace
                    
                    self.update_log.append({
                        "timestamp": datetime.now().isoformat(),
                        "operation": "add_sheet_replaced" if was_replaced else "add_sheet",
                        "cell": f"{sheet_name} → {target_sheet_name}",
                        "status": "success",
                        "details": f"{'Replaced' if was_replaced else 'Added'} sheet '{target_sheet_name}' from source sheet '{sheet_name}'"
                    })
                
                self._log_info(f"Successfully added {len(sheets_to_copy)} sheets")
                
            finally:
                # Always close source workbook
                source_workbook.close()
            
            # Add diagnostic log sheet if requested
            if include_update_log:
                self._add_update_log_sheet()
            
            # Save updated workbook
            output_path = self._save_updated_workbook()
            self._log_info(f"Sheet addition completed successfully: {output_path}")
            
            return output_path
            
        except Exception as e:
            self._log_error(f"Sheet addition failed: {e}")
            raise ExcelProcessingError(f"Failed to add sheets: {e}")

    def _insert_image_v2(
        self, sheet: Worksheet, cell_address: str, image_data: Any
    ) -> bool:
        """Insert image using direct file approach to avoid PIL corruption."""
        try:
            if not image_data:
                sheet[cell_address].value = "[NO_IMAGE_DATA]"
                self._log_warning(f"No image data provided for {cell_address}")
                return True

            # Determine final image path - avoid PIL processing that causes corruption
            final_image_path = None
            original_size = None

            if isinstance(image_data, str):
                if image_data.startswith("file://") or os.path.exists(image_data):
                    # Local file path - use directly when possible
                    try:
                        file_path = image_data
                        if file_path.startswith("file://"):
                            import urllib.parse

                            file_path = urllib.parse.unquote(file_path[7:])

                        if os.path.exists(file_path):
                            # Get original size without processing
                            with open(file_path, "rb") as f:
                                image_bytes = f.read()
                            temp_img = PILImage.open(io.BytesIO(image_bytes))
                            original_size = temp_img.size
                            temp_img.close()

                            # Use original file directly - no processing to avoid corruption
                            final_image_path = file_path
                            self._log_info(
                                f"Using original file directly: {file_path} (size: {original_size[0]}x{original_size[1]})"
                            )
                        else:
                            raise ValueError(f"File not found: {file_path}")
                    except Exception as e:
                        self._log_error(
                            f"Failed to process local file {image_data}: {e}"
                        )
                        sheet[cell_address].value = "!#VALUE"
                        return False

                elif image_data.startswith("data:image") or image_data.startswith(
                    "http"
                ):
                    # For base64 and URLs, save to temp file without resizing
                    try:
                        if image_data.startswith("data:image"):
                            image_bytes = self._decode_base64_image(image_data)
                        else:
                            image_bytes = self._download_image(image_data)

                        # Get size info
                        temp_img = PILImage.open(io.BytesIO(image_bytes))
                        original_size = temp_img.size
                        temp_img.close()

                        # Save raw bytes to temp file (no PIL processing)
                        import tempfile

                        temp_fd, temp_image_path = tempfile.mkstemp(suffix=".png")
                        os.close(temp_fd)
                        with open(temp_image_path, "wb") as f:
                            f.write(image_bytes)

                        self.temp_image_files.append(temp_image_path)
                        final_image_path = temp_image_path
                        self._log_info(
                            f"Saved to temp file: {temp_image_path} (size: {original_size[0]}x{original_size[1]})"
                        )
                    except Exception as e:
                        self._log_error(f"Failed to process image data: {e}")
                        sheet[cell_address].value = "!#VALUE"
                        return False
                else:
                    # Text fallback
                    sheet[cell_address].value = str(image_data)
                    self._log_info(
                        f"Inserted text instead of image at {cell_address}: {image_data}"
                    )
                    return True
            else:
                # Unknown format - use text
                sheet[cell_address].value = str(image_data)
                self._log_warning(
                    f"Unknown image format, inserted as text at {cell_address}"
                )
                return True

            if not final_image_path or not original_size:
                sheet[cell_address].value = "!#VALUE"
                self._log_error(f"Could not process image data for {cell_address}")
                return False

            # Create Excel image from final path (avoiding PIL altogether)
            excel_img = ExcelImage(final_image_path)

            # Apply width-constrained resizing with aspect ratio preservation
            target_width_points = 200  # Target width in points
            target_width_pixels = target_width_points * 1.33  # Convert to pixels

            # Calculate proper height based on original aspect ratio
            original_width, original_height = original_size
            aspect_ratio = original_height / original_width
            target_height_pixels = target_width_pixels * aspect_ratio

            # Set both width and height explicitly to maintain aspect ratio
            excel_img.width = target_width_pixels
            excel_img.height = target_height_pixels

            # Position at target cell
            row, col = self._parse_cell_address(cell_address)
            excel_img.anchor = f"{get_column_letter(col)}{row}"

            sheet.add_image(excel_img)
            # Explicitly ensure cell remains empty - no text pollution
            sheet[cell_address].value = None

            # Adjust row height based on new calculated height
            self._adjust_row_height(sheet, row, target_height_pixels)

            self._log_success(
                f"Inserted image at {cell_address} (original: {original_size[0]}x{original_size[1]}, resized: {target_width_pixels:.0f}x{target_height_pixels:.0f}px)"
            )
            return True

        except Exception as e:
            self._log_error(f"Image insert failed at {cell_address}: {e}")
            try:
                # Try to set error value if valid cell address
                if self._is_cell_address(cell_address):
                    sheet[cell_address].value = "!#VALUE"
            except:
                pass  # If even error setting fails, just continue
            return False
