"""Excel range to image exporter using local table recreation."""

import logging
import os
import tempfile
import time
import io
import base64
from typing import Dict, Any, Optional, List, Tuple
from dataclasses import dataclass
from pathlib import Path

import matplotlib.pyplot as plt
import matplotlib.patches as patches
from matplotlib.table import Table
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill
from PIL import Image, ImageDraw, ImageFont

from .temp_file_manager import TempFileManager
from .utils.exceptions import ExcelProcessingError, ValidationError

logger = logging.getLogger(__name__)


@dataclass
class RangeImageConfig:
    """Configuration for range image export."""

    field_name: str
    sheet_name: str
    range: str
    include_headers: bool = True
    output_format: str = "png"
    dpi: int = 150
    fit_to_content: bool = True
    width: Optional[int] = None
    height: Optional[int] = None


@dataclass
class RangeImageResult:
    """Result of range image export."""

    field_name: str
    image_path: str
    image_data: bytes
    width: int
    height: int
    range_dimensions: Tuple[int, int]  # (rows, columns)
    success: bool
    error_message: Optional[str] = None


class ExcelRangeExporter:
    """Exports Excel ranges as images using local table recreation."""

    def __init__(self):
        """Initialize the Excel range exporter."""
        self.temp_manager = TempFileManager()
        self.debug_directory = None

    def set_debug_directory(self, debug_dir: str) -> None:
        """Set debug directory for saving range images in development mode."""
        self.debug_directory = debug_dir
        logger.info(f"Debug directory set for range images: {debug_dir}")

    def export_ranges(
        self, workbook_path: str, range_configs: List[RangeImageConfig]
    ) -> List[RangeImageResult]:
        """Export multiple ranges from an Excel workbook using local table recreation."""
        if not os.path.exists(workbook_path):
            raise ExcelProcessingError(f"Workbook not found: {workbook_path}")

        if not range_configs:
            logger.warning("No range configurations provided")
            return []

        results = []

        try:
            # Load workbook locally
            logger.info(f"Loading Excel workbook: {workbook_path}")
            wb = load_workbook(workbook_path, data_only=True)
            
            # Get available worksheets for validation
            worksheets = wb.sheetnames
            logger.info(f"Available worksheets: {worksheets}")

            # Process each range configuration
            for config in range_configs:
                result = self._export_single_range_local(wb, config, worksheets)
                results.append(result)

        except Exception as e:
            logger.error(f"Error during range export: {e}")
            raise ExcelProcessingError(f"Failed to export ranges: {e}")

        return results

    def _export_single_range_local(
        self, workbook, config: RangeImageConfig, available_worksheets: List[str]
    ) -> RangeImageResult:
        """Export a single range as image using local table recreation."""
        try:
            # Validate sheet exists
            if config.sheet_name not in available_worksheets:
                error_msg = f"Sheet '{config.sheet_name}' not found. Available: {available_worksheets}"
                logger.error(error_msg)
                return RangeImageResult(
                    field_name=config.field_name,
                    image_path="",
                    image_data=b"",
                    width=0,
                    height=0,
                    range_dimensions=(0, 0),
                    success=False,
                    error_message=error_msg,
                )

            # Get worksheet
            worksheet = workbook[config.sheet_name]
            
            # Parse range
            range_cells = worksheet[config.range]
            
            # Create table image using matplotlib
            logger.info(
                f"Creating table image for range {config.range} from sheet '{config.sheet_name}'"
            )
            image_data = self._create_table_image(range_cells, config)

            # Save image to temporary file
            image_path = self._save_image_to_temp_file(
                image_data, config.field_name, config.output_format
            )

            # Get actual image dimensions
            actual_width, actual_height = self._get_image_dimensions(image_data)

            # Calculate range dimensions
            if hasattr(range_cells, '__iter__') and hasattr(range_cells[0], '__iter__'):
                rows = len(range_cells)
                cols = len(range_cells[0]) if rows > 0 else 0
                range_dimensions = (rows, cols)
            else:
                range_dimensions = (1, 1)

            logger.info(f"Successfully exported range {config.range} to {image_path}")
            return RangeImageResult(
                field_name=config.field_name,
                image_path=image_path,
                image_data=image_data,
                width=actual_width,
                height=actual_height,
                range_dimensions=range_dimensions,
                success=True,
            )

        except Exception as e:
            error_msg = f"Unexpected error exporting range {config.range}: {e}"
            logger.error(error_msg)
            return RangeImageResult(
                field_name=config.field_name,
                image_path="",
                image_data=b"",
                width=0,
                height=0,
                range_dimensions=(0, 0),
                success=False,
                error_message=error_msg,
            )

    def _create_table_image(self, range_cells, config: RangeImageConfig) -> bytes:
        """Create a table image from Excel range cells using matplotlib."""
        
        # Extract data and formatting
        table_data = []
        cell_colors = []
        text_colors = []
        font_weights = []
        
        # Handle both single cell and multi-cell ranges
        if not hasattr(range_cells, '__iter__') or not hasattr(range_cells[0], '__iter__'):
            # Single cell
            range_cells = [[range_cells]]
        
        for row in range_cells:
            row_data = []
            row_colors = []
            row_text_colors = []
            row_weights = []
            
            for cell in row:
                # Get cell value
                value = cell.value
                if value is None:
                    value = ""
                elif isinstance(value, (int, float)):
                    if cell.number_format and '£' in cell.number_format:
                        value = f"£{value:,.2f}" if isinstance(value, float) else f"£{value:,}"
                    else:
                        value = str(value)
                else:
                    value = str(value)
                
                row_data.append(value)
                
                # Get background color
                if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                    try:
                        hex_color = cell.fill.start_color.rgb
                        if len(hex_color) == 8:  # ARGB format
                            hex_color = hex_color[2:]  # Remove alpha
                        rgb = tuple(int(hex_color[i:i+2], 16)/255.0 for i in (0, 2, 4))
                        row_colors.append(rgb)
                    except:
                        row_colors.append('white')
                else:
                    row_colors.append('white')
                
                # Get text color
                if cell.font and cell.font.color and cell.font.color.rgb:
                    try:
                        hex_color = cell.font.color.rgb
                        if len(hex_color) == 8:  # ARGB format
                            hex_color = hex_color[2:]  # Remove alpha
                        rgb = tuple(int(hex_color[i:i+2], 16)/255.0 for i in (0, 2, 4))
                        row_text_colors.append(rgb)
                    except:
                        row_text_colors.append('black')
                else:
                    row_text_colors.append('black')
                
                # Get font weight
                if cell.font and cell.font.bold:
                    row_weights.append('bold')
                else:
                    row_weights.append('normal')
            
            table_data.append(row_data)
            cell_colors.append(row_colors)
            text_colors.append(row_text_colors)
            font_weights.append(row_weights)
        
        # Create matplotlib figure
        plt.style.use('default')
        fig, ax = plt.subplots(figsize=(16, max(6, len(table_data) * 0.5)))
        ax.axis('tight')
        ax.axis('off')
        
        # Create table
        table = ax.table(
            cellText=table_data,
            cellLoc='center',
            loc='center',
            colWidths=[0.12] * len(table_data[0]) if table_data else [0.12]
        )
        
        # Style the table
        table.auto_set_font_size(False)
        table.set_fontsize(9)
        table.scale(1, 2)  # Make cells taller
        
        # Apply formatting
        for (i, j), cell in table.get_celld().items():
            if i < len(cell_colors) and j < len(cell_colors[i]):
                # Set background color
                cell.set_facecolor(cell_colors[i][j])
                
                # Set text color and weight
                cell.set_text_props(
                    color=text_colors[i][j],
                    weight=font_weights[i][j]
                )
                
                # Add border
                cell.set_edgecolor('lightgray')
                cell.set_linewidth(0.5)
        
        # Adjust layout
        plt.tight_layout()
        plt.subplots_adjust(left=0.05, right=0.95, top=0.95, bottom=0.05)
        
        # Save to bytes
        img_buffer = io.BytesIO()
        plt.savefig(
            img_buffer, 
            format='png', 
            dpi=config.dpi or 150,
            bbox_inches='tight',
            facecolor='white',
            edgecolor='none'
        )
        plt.close()  # Important: close the figure to free memory
        
        img_buffer.seek(0)
        return img_buffer.getvalue()

    def _save_image_to_temp_file(
        self, image_data: bytes, field_name: str, output_format: str
    ) -> str:
        """Save image data to temporary file."""
        # Generate unique filename
        timestamp = int(time.time())
        filename = f"range_image_{field_name}_{timestamp}.{output_format.lower()}"

        # If debug directory is set, save there instead of temp directory
        if self.debug_directory:
            try:
                os.makedirs(self.debug_directory, exist_ok=True)
                temp_path = os.path.join(self.debug_directory, filename)

                with open(temp_path, "wb") as f:
                    f.write(image_data)

                logger.info(f"Saved range image to debug directory: {temp_path}")
                return temp_path
            except Exception as e:
                logger.warning(
                    f"Failed to save to debug directory, falling back to temp: {e}"
                )
                # Fall back to temp file if debug save fails

        # Use basic temp file for now (can be improved later)
        import tempfile
        temp_fd, temp_path = tempfile.mkstemp(suffix=f".{output_format.lower()}")
        os.close(temp_fd)  # Close the file descriptor since we'll use the path

        try:
            with open(temp_path, "wb") as f:
                f.write(image_data)

            logger.debug(f"Saved range image to temporary file: {temp_path}")
            return temp_path

        except Exception as e:
            logger.error(f"Failed to save image to temporary file: {e}")
            raise ExcelProcessingError(f"Failed to save image: {e}")

    def _get_image_dimensions(self, image_data: bytes) -> Tuple[int, int]:
        """Get image dimensions from binary data."""
        try:
            from PIL import Image
            import io

            image = Image.open(io.BytesIO(image_data))
            return image.size  # (width, height)

        except Exception as e:
            logger.warning(f"Failed to get image dimensions: {e}")
            return (0, 0)

    def validate_config(self, config: RangeImageConfig) -> bool:
        """Validate a range image configuration."""
        try:
            # Check required fields
            if not config.field_name:
                raise ValidationError("field_name is required")
            if not config.sheet_name:
                raise ValidationError("sheet_name is required")
            if not config.range:
                raise ValidationError("range is required")

            # Validate range format (basic check)
            if not self._is_valid_range_format(config.range):
                raise ValidationError(f"Invalid range format: {config.range}")

            # Validate output format
            valid_formats = ["png", "jpg", "jpeg"]
            if config.output_format.lower() not in valid_formats:
                raise ValidationError(f"Invalid output format: {config.output_format}")

            # Validate DPI
            if config.dpi < 72 or config.dpi > 600:
                raise ValidationError(
                    f"DPI must be between 72 and 600, got: {config.dpi}"
                )

            return True

        except ValidationError:
            raise
        except Exception as e:
            raise ValidationError(f"Configuration validation failed: {e}")

    def _is_valid_range_format(self, range_str: str) -> bool:
        """Check if range string has valid Excel format (e.g., A1:E15)."""
        import re

        # Basic pattern for Excel range: Letter(s) + Number + : + Letter(s) + Number
        pattern = r"^[A-Z]+\d+:[A-Z]+\d+$"
        return bool(re.match(pattern, range_str.upper()))

    def cleanup_temp_files(self) -> None:
        """Clean up temporary files created during export."""
        try:
            self.temp_manager.cleanup_all()
            logger.info("Cleaned up temporary range image files")
        except Exception as e:
            logger.warning(f"Failed to cleanup temporary files: {e}")


def create_range_configs_from_dict(
    range_configs_data: List[Dict[str, Any]],
) -> List[RangeImageConfig]:
    """Create RangeImageConfig objects from dictionary data."""
    configs = []

    for config_data in range_configs_data:
        try:
            config = RangeImageConfig(**config_data)
            configs.append(config)
        except TypeError as e:
            logger.error(f"Invalid range configuration: {config_data}, error: {e}")
            raise ValidationError(f"Invalid range configuration: {e}")

    return configs


def validate_range_configs(configs: List[RangeImageConfig]) -> List[str]:
    """Validate multiple range configurations and return list of errors."""
    errors = []
    field_names = set()

    for i, config in enumerate(configs):
        try:
            # Check for duplicate field names
            if config.field_name in field_names:
                errors.append(f"Config {i}: Duplicate field_name '{config.field_name}'")
            else:
                field_names.add(config.field_name)

            # Validate individual config (this will raise ValidationError if invalid)
            exporter = ExcelRangeExporter()  # Dummy exporter for validation
            exporter.validate_config(config)

        except ValidationError as e:
            errors.append(f"Config {i}: {e}")
        except Exception as e:
            errors.append(f"Config {i}: Unexpected error - {e}")

    return errors
