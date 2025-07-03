"""Basic functional tests for ExcelUpdater module."""

import os
import tempfile
import json
from unittest import TestCase
from openpyxl import Workbook
from src.excel_updater import ExcelUpdater


class TestExcelUpdater(TestCase):
    """Test ExcelUpdater functionality."""
    
    def setUp(self):
        """Set up test fixtures."""
        # Create temporary Excel file
        self.temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        self.temp_file.close()
        
        # Create test workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Order Form"
        
        # Add test data structure
        # Client info section
        ws['A2'] = 'Client'
        ws['B2'] = 'Test Client'
        ws['C2'] = 'Word Or Image'
        ws['D2'] = 'Word'
        
        # Search criteria table
        ws['A5'] = 'Word'
        ws['B5'] = 'Search Criteria'
        ws['C5'] = 'Remarks'
        ws['A6'] = 'ExistingWord'
        ws['B6'] = 'Exact'
        ws['C6'] = 'Old remark'
        
        # Image search table
        ws['A10'] = 'Image'
        ws['B10'] = 'Search Criteria'
        ws['C10'] = 'Image Class.Division.Subdivision'
        
        wb.save(self.temp_file.name)
        wb.close()
        
    def tearDown(self):
        """Clean up test fixtures."""
        if os.path.exists(self.temp_file.name):
            os.unlink(self.temp_file.name)
    
    def test_cell_address_update(self):
        """Test direct cell address updates."""
        config = {
            "sheet_configs": {
                "Order Form": {
                    "subtables": [
                        {
                            "name": "test_data",
                            "type": "key_value_pairs",
                            "header_search": {
                                "method": "cell_address",
                                "cell": "B2"
                            },
                            "data_update": {
                                "column_mappings": {
                                    "B2": {
                                        "name": "test_value",
                                        "type": "text"
                                    }
                                }
                            }
                        }
                    ]
                }
            }
        }
        
        update_data = {
            "test_data": {
                "test_value": "Updated Value"
            }
        }
        
        updater = ExcelUpdater(self.temp_file.name)
        result_path = updater.update_excel(update_data, config)
        updater.close()
        
        # Verify file was created
        self.assertTrue(os.path.exists(result_path))
        
        # Clean up result file
        os.unlink(result_path)
    
    def test_contains_text_update(self):
        """Test text search method updates."""
        config = {
            "sheet_configs": {
                "Order Form": {
                    "subtables": [
                        {
                            "name": "client_info",
                            "type": "key_value_pairs", 
                            "header_search": {
                                "method": "contains_text",
                                "text": "Client",
                                "column": "A",
                                "search_range": "A1:A10"
                            },
                            "data_update": {
                                "orientation": "horizontal",
                                "headers_row_offset": 0,
                                "data_row_offset": 0,
                                "column_mappings": {
                                    "Client": {
                                        "name": "client_name",
                                        "type": "text"
                                    }
                                }
                            }
                        }
                    ]
                }
            }
        }
        
        update_data = {
            "client_info": {
                "client_name": "New Client Name"
            }
        }
        
        updater = ExcelUpdater(self.temp_file.name)
        result_path = updater.update_excel(update_data, config)
        updater.close()
        
        # Verify file was created
        self.assertTrue(os.path.exists(result_path))
        
        # Clean up result file
        os.unlink(result_path)
    
    def test_table_update(self):
        """Test table data updates."""
        config = {
            "sheet_configs": {
                "Order Form": {
                    "subtables": [
                        {
                            "name": "search_criteria",
                            "type": "table",
                            "header_search": {
                                "method": "contains_text",
                                "text": "Word",
                                "column": "A", 
                                "search_range": "A3:A10"
                            },
                            "data_update": {
                                "orientation": "vertical",
                                "headers_row_offset": 0,
                                "data_row_offset": 1,
                                "column_mappings": {
                                    "Word": {
                                        "name": "word",
                                        "type": "text"
                                    },
                                    "Search Criteria": {
                                        "name": "search_criteria", 
                                        "type": "text"
                                    }
                                }
                            }
                        }
                    ]
                }
            }
        }
        
        update_data = {
            "search_criteria": [
                {
                    "word": "NewWord1",
                    "search_criteria": "Exact Match"
                },
                {
                    "word": "NewWord2", 
                    "search_criteria": "Similar"
                }
            ]
        }
        
        updater = ExcelUpdater(self.temp_file.name)
        result_path = updater.update_excel(update_data, config)
        updater.close()
        
        # Verify file was created
        self.assertTrue(os.path.exists(result_path))
        
        # Clean up result file
        os.unlink(result_path)
    
    def test_image_update_base64(self):
        """Test base64 image updates."""
        config = {
            "sheet_configs": {
                "Order Form": {
                    "subtables": [
                        {
                            "name": "image_test",
                            "type": "key_value_pairs",
                            "header_search": {
                                "method": "cell_address",
                                "cell": "A10"
                            },
                            "data_update": {
                                "column_mappings": {
                                    "A11": {
                                        "name": "test_image",
                                        "type": "image"
                                    }
                                }
                            }
                        }
                    ]
                }
            }
        }
        
        # Small test image (1x1 transparent PNG)
        base64_image = "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mNk+M9QDwADhgGAWjR9awAAAABJRU5ErkJggg=="
        
        update_data = {
            "image_test": {
                "test_image": base64_image
            }
        }
        
        updater = ExcelUpdater(self.temp_file.name)
        result_path = updater.update_excel(update_data, config)
        updater.close()
        
        # Verify file was created
        self.assertTrue(os.path.exists(result_path))
        
        # Clean up result file
        os.unlink(result_path)
    
    def test_validation_error(self):
        """Test validation error handling."""
        config = {
            "sheet_configs": {
                "Order Form": {
                    "subtables": [
                        {
                            "name": "invalid_config",
                            "type": "key_value_pairs",
                            "header_search": {
                                "method": "invalid_method"
                            }
                        }
                    ]
                }
            }
        }
        
        update_data = {"invalid_config": {"test": "value"}}
        
        updater = ExcelUpdater(self.temp_file.name)
        
        with self.assertRaises(Exception):
            updater.update_excel(update_data, config)
        
        updater.close()
    
    def test_missing_sheet_handling(self):
        """Test handling of missing sheets."""
        config = {
            "sheet_configs": {
                "NonExistent Sheet": {
                    "subtables": [
                        {
                            "name": "test_data",
                            "type": "key_value_pairs",
                            "header_search": {
                                "method": "cell_address",
                                "cell": "A1"
                            },
                            "data_update": {
                                "column_mappings": {
                                    "A1": {
                                        "name": "test_value",
                                        "type": "text"
                                    }
                                }
                            }
                        }
                    ]
                }
            }
        }
        
        update_data = {"test_data": {"test_value": "test"}}
        
        updater = ExcelUpdater(self.temp_file.name)
        # Should not raise error, but should log warning
        result_path = updater.update_excel(update_data, config)
        updater.close()
        
        # Verify file was still created
        self.assertTrue(os.path.exists(result_path))
        
        # Clean up result file
        os.unlink(result_path)