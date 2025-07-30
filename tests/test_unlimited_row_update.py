"""Test unlimited row update functionality."""

import pytest
import os
import tempfile
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from src.excel_updater import ExcelUpdater


def create_test_excel_with_content_below():
    """Create a test Excel file with a table and content below it."""
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    
    # Create headers for a table at row 5
    ws["A5"] = "Product Table"
    ws["A6"] = "ID"
    ws["B6"] = "Name"
    ws["C6"] = "Price"
    
    # Add initial data (3 rows)
    initial_data = [
        (1, "Product A", 100),
        (2, "Product B", 200),
        (3, "Product C", 300),
    ]
    
    for idx, (id_val, name, price) in enumerate(initial_data, start=7):
        ws[f"A{idx}"] = id_val
        ws[f"B{idx}"] = name
        ws[f"C{idx}"] = price
    
    # Add some empty rows (table boundary)
    # Rows 10-11 are empty
    
    # Add content below the table that should be preserved
    ws["A13"] = "Summary Section"
    ws["B13"] = "Total Products:"
    ws["C13"] = "=COUNTA(B7:B9)"  # Formula counting products
    
    ws["A14"] = "Average Price:"
    ws["B14"] = "=AVERAGE(C7:C9)"  # Formula for average
    
    # Add a merged cell
    ws.merge_cells("A16:C16")
    ws["A16"] = "This is a merged cell that should move down"
    
    # Add some formatting
    ws["A13"].font = ws["A13"].font.copy(bold=True)
    
    # Create temporary file
    temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(temp_file.name)
    temp_file.close()
    
    return temp_file.name


def test_unlimited_row_expansion():
    """Test that tables can expand beyond their original size."""
    # Create test file
    test_file = create_test_excel_with_content_below()
    
    try:
        # Configuration for the update
        config = {
            "sheet_configs": {
                "TestSheet": {
                    "subtables": [
                        {
                            "name": "products",
                            "type": "table",
                            "header_search": {
                                "method": "contains_text",
                                "text": "Product Table",
                                "column": "A",
                                "search_range": "A1:A10"
                            },
                            "data_update": {
                                "headers_row_offset": 1,
                                "data_row_offset": 1,
                                "column_mappings": {
                                    "ID": {"name": "id", "type": "number"},
                                    "Name": {"name": "name", "type": "text"},
                                    "Price": {"name": "price", "type": "number"}
                                },
                                "max_expansion_rows": 1000,
                                "expansion_behavior": "preserve_below"
                            }
                        }
                    ]
                }
            }
        }
        
        # New data with 7 rows (expansion of 4 rows)
        update_data = {
            "products": [
                {"id": 1, "name": "Updated Product A", "price": 150},
                {"id": 2, "name": "Updated Product B", "price": 250},
                {"id": 3, "name": "Updated Product C", "price": 350},
                {"id": 4, "name": "New Product D", "price": 400},
                {"id": 5, "name": "New Product E", "price": 500},
                {"id": 6, "name": "New Product F", "price": 600},
                {"id": 7, "name": "New Product G", "price": 700},
            ]
        }
        
        # Update the Excel file
        updater = ExcelUpdater(test_file)
        output_file = updater.update_excel(update_data, config)
        updater.close()
        
        # Verify the results
        wb = load_workbook(output_file)
        ws = wb.active
        
        # Check that all 7 products are present
        for idx, product in enumerate(update_data["products"], start=7):
            assert ws[f"A{idx}"].value == product["id"]
            assert ws[f"B{idx}"].value == product["name"]
            assert ws[f"C{idx}"].value == product["price"]
        
        # Check that content below was preserved and shifted down by 4 rows
        assert ws["A17"].value == "Summary Section"  # Was at row 13, now at 17 (13+4)
        assert ws["B17"].value == "Total Products:"
        
        # Check that formulas were preserved and moved (not expanded)
        # The original formula =COUNTA(B7:B9) should still reference the same range
        assert ws["C17"].value is not None  # Formula should exist
        assert ws["B18"].value is not None  # Average formula should exist
        
        # Check merged cell was preserved and moved
        assert "This is a merged cell" in str(ws["A20"].value)  # Was at 16, now at 20
        
        # Cleanup
        os.unlink(test_file)
        os.unlink(output_file)
        
    except Exception as e:
        # Cleanup on error
        if os.path.exists(test_file):
            os.unlink(test_file)
        raise e


def test_multiple_tables_with_expansion():
    """Test multiple tables where one expands and affects the position of another."""
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    
    # First table at row 2
    ws["A2"] = "Table 1"
    ws["A3"] = "Key"
    ws["B3"] = "Value"
    
    # Fixed-size key-value pairs
    ws["A4"] = "Setting1"
    ws["B4"] = "Value1"
    ws["A5"] = "Setting2"
    ws["B5"] = "Value2"
    
    # Second table at row 8 (expandable)
    ws["A8"] = "Table 2"
    ws["A9"] = "Item"
    ws["B9"] = "Count"
    
    # Initial data (2 rows)
    ws["A10"] = "Item1"
    ws["B10"] = 10
    ws["A11"] = "Item2"
    ws["B11"] = 20
    
    # Save test file
    temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(temp_file.name)
    temp_file.close()
    
    try:
        # Configuration with two tables
        config = {
            "sheet_configs": {
                "TestSheet": {
                    "subtables": [
                        {
                            "name": "settings",
                            "type": "key_value_pairs",
                            "header_search": {
                                "method": "contains_text",
                                "text": "Table 1",
                                "column": "A",
                                "search_range": "A1:A5"
                            },
                            "data_update": {
                                "headers_row_offset": 1,
                                "data_row_offset": 1,
                                "orientation": "vertical",
                                "data_col_offset": 1,
                                "column_mappings": {
                                    "Setting1": {"name": "setting1", "type": "text"},
                                    "Setting2": {"name": "setting2", "type": "text"}
                                }
                            }
                        },
                        {
                            "name": "items",
                            "type": "table",
                            "header_search": {
                                "method": "contains_text",
                                "text": "Table 2",
                                "column": "A",
                                "search_range": "A6:A10"
                            },
                            "data_update": {
                                "headers_row_offset": 1,
                                "data_row_offset": 1,
                                "column_mappings": {
                                    "Item": {"name": "item", "type": "text"},
                                    "Count": {"name": "count", "type": "number"}
                                },
                                "max_expansion_rows": 100
                            }
                        }
                    ]
                }
            }
        }
        
        # Update data - settings stay same, items expand to 5 rows
        update_data = {
            "settings": {
                "setting1": "Updated Value 1",
                "setting2": "Updated Value 2"
            },
            "items": [
                {"item": "Item1", "count": 15},
                {"item": "Item2", "count": 25},
                {"item": "Item3", "count": 35},
                {"item": "Item4", "count": 45},
                {"item": "Item5", "count": 55},
            ]
        }
        
        # Update the Excel file
        updater = ExcelUpdater(temp_file.name)
        output_file = updater.update_excel(update_data, config)
        updater.close()
        
        # Verify results
        wb = load_workbook(output_file)
        ws = wb.active
        
        # Check fixed table (should not move)
        assert ws["B4"].value == "Updated Value 1"
        assert ws["B5"].value == "Updated Value 2"
        
        # Check expandable table (all 5 items)
        for idx, item in enumerate(update_data["items"], start=10):
            assert ws[f"A{idx}"].value == item["item"]
            assert ws[f"B{idx}"].value == item["count"]
        
        # Cleanup
        os.unlink(temp_file.name)
        os.unlink(output_file)
        
    except Exception as e:
        # Cleanup on error
        if os.path.exists(temp_file.name):
            os.unlink(temp_file.name)
        raise e


if __name__ == "__main__":
    test_unlimited_row_expansion()
    test_multiple_tables_with_expansion()
    print("All tests passed!")