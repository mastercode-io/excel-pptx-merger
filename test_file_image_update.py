#!/usr/bin/env python3
"""Test script for file path image update."""

import json
import requests
import tempfile
import os
from openpyxl import Workbook

def create_test_excel_file():
    """Create test Excel file for API testing.""" 
    # Create unique temp file each time to avoid corruption
    import time
    timestamp = int(time.time())
    temp_file_path = f'/tmp/test_excel_{timestamp}.xlsx'
    
    # Create test workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Order Form"
    
    # Image search table
    ws['A15'] = 'Image'
    ws['B15'] = 'Search Criteria'
    ws['C15'] = 'Image Class.Division.Subdivision'
    
    wb.save(temp_file_path)
    wb.close()
    
    return temp_file_path

def test_file_image_update():
    """Test file path image update."""
    # Create test files
    excel_file_path = create_test_excel_file()
    config_file_path = "/Users/alexsherin/Projects_/excel-pptx-merger/example_update_config.json"
    data_file_path = "/Users/alexsherin/Projects_/excel-pptx-merger/test_file_image_data.json"
    
    try:
        # Read configuration and data
        with open(config_file_path, 'r') as f:
            config = json.load(f)
        
        with open(data_file_path, 'r') as f:
            update_data = json.load(f)
        
        # Only test image_search subtable
        filtered_config = {
            "version": "1.0",
            "sheet_configs": {
                "Order Form": {
                    "subtables": [subtable for subtable in config["sheet_configs"]["Order Form"]["subtables"] 
                                 if subtable["name"] == "image_search"]
                }
            }
        }
        
        # Prepare API request
        url = "http://localhost:8080/api/v1/update"
        
        with open(excel_file_path, 'rb') as excel_file:
            files = {
                'excel_file': ('test_file_image.xlsx', excel_file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            }
            
            data = {
                'config': json.dumps(filtered_config),
                'update_data': json.dumps(update_data)
            }
            
            print("Testing file path image update...")
            print("Image sources:")
            for img in update_data["image_search"]:
                print(f"  - {img['image'][:50]}...")
            
            # Make request
            response = requests.post(url, files=files, data=data)
            
            if response.status_code == 200:
                # Save response file
                output_path = "updated_file_image_test.xlsx"
                with open(output_path, 'wb') as f:
                    f.write(response.content)
                print(f"✅ Success! Updated file saved as: {output_path}")
                return True
            else:
                print(f"❌ Error {response.status_code}: {response.text}")
                return False
    
    except Exception as e:
        print(f"❌ Test failed: {e}")
        return False
    finally:
        # Cleanup temp file
        if os.path.exists(excel_file_path):
            os.unlink(excel_file_path)

if __name__ == "__main__":
    success = test_file_image_update()
    exit(0 if success else 1)